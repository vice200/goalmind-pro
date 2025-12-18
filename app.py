import os
from io import BytesIO, StringIO
from typing import Dict, List, Tuple

import pandas as pd
import numpy as np
import requests
import joblib
import html
import textwrap
import streamlit as st
from math import exp, factorial

from sklearn.ensemble import RandomForestClassifier, RandomForestRegressor
import datetime
from typing import Dict, List, Tuple  # ako već nije gore

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule
import glob


# =========================
# RECO / VALUE helpers
# =========================

def implied_prob_1x2_norm(odd_h, odd_d, odd_a):
    qh, qd, qa = 1/odd_h, 1/odd_d, 1/odd_a
    s = qh + qd + qa
    if s <= 0:
        return None, None, None
    return qh/s, qd/s, qa/s

def kelly_fraction(p, odds):
    if p is None or odds is None or odds <= 1:
        return 0.0
    k = (p * odds - 1.0) / (odds - 1.0)
    return max(0.0, k)

def p_blend(p_poi, p_ai, w_ai=0.55):
    if p_poi is None and p_ai is None:
        return None
    if p_poi is None:
        return p_ai
    if p_ai is None:
        return p_poi
    return w_ai*p_ai + (1-w_ai)*p_poi

MIN_ABS_EDGE_1X2 = 0.03   # +3pp (ubija fake value longshot)
MIN_KELLY_1X2    = 0.01   # >= 1% banke
MAX_DISAGREE     = 0.18   # ako AI i Poi previše različiti -> ne može HIGH

def pick_1x2_reco(row, w_ai=0.55):
    oh, od, oa = row.get("odds_h"), row.get("odds_d"), row.get("odds_a")
    if not (oh and od and oa):
        return ("No bet", "NONE", None, None, None)

    qh, qd, qa = implied_prob_1x2_norm(oh, od, oa)
    if qh is None:
        return ("No bet", "NONE", None, None, None)

    pH_poi, pD_poi, pA_poi = row.get("p_poi_h"), row.get("p_poi_d"), row.get("p_poi_a")
    pH_ai,  pD_ai,  pA_ai  = row.get("p_ai_h"),  row.get("p_ai_d"),  row.get("p_ai_a")

    pH = p_blend(pH_poi, pH_ai, w_ai=w_ai)
    pD = p_blend(pD_poi, pD_ai, w_ai=w_ai)
    pA = p_blend(pA_poi, pA_ai, w_ai=w_ai)

    # max neslaganje po ishodu
    dis = 0.0
    for a, b in [(pH_poi, pH_ai), (pD_poi, pD_ai), (pA_poi, pA_ai)]:
        if a is not None and b is not None:
            dis = max(dis, abs(a - b))

    cands = [
        ("Home win (1)", pH, oh, qh),
        ("Draw (X)",     pD, od, qd),
        ("Away win (2)", pA, oa, qa),
    ]

    best = None
    for name, p, odds, q in cands:
        if p is None:
            continue
        abs_edge = p - q
        rel_edge = abs_edge / q if q > 0 else 0.0
        kelly = kelly_fraction(p, odds)

        if abs_edge < MIN_ABS_EDGE_1X2:
            continue
        if kelly < MIN_KELLY_1X2:
            continue

        ev = (p * odds - 1.0)  # expected value
        item = (ev, name, rel_edge, abs_edge, kelly, dis)
        if best is None or item[0] > best[0]:
            best = item

    if best is None:
        return ("No bet", "NONE", None, None, None)

    _, name, rel_edge, abs_edge, kelly, dis = best

    if dis > MAX_DISAGREE:
        conf = "MEDIUM" if abs_edge >= 0.04 else "LOW"
    else:
        conf = "HIGH" if abs_edge >= 0.05 else ("MEDIUM" if abs_edge >= 0.04 else "LOW")

    return (name, conf, rel_edge, abs_edge, kelly)

# ============================
# TEAM NAME NORMALIZATION + MAPPING
# ============================

def inject_pro_css():
    st.markdown("""
    <style>
    /* MAIN APP AREA */
    .block-container {
        padding-top: 1.5rem;
        padding-bottom: 2rem;
        max-width: 1350px;
    }

    /* Background subtle tweak */
    [data-testid="stAppViewContainer"] {
        background: radial-gradient(circle at top left, #020617 0, #020617 40%, #000000 100%);
    }

    /* HERO HEADER */
    .hero {
        padding: 1.3rem 1.6rem;
        border-radius: 18px;
        background: linear-gradient(135deg, #0f172a, #1e293b 60%);
        color: #e5e7eb;
        margin-bottom: 1.3rem;
        border: 1px solid #1f2937;
        box-shadow: 0 0 18px rgba(0,0,0,0.35);
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 1.5rem;
    }
    .hero-left {
        flex: 3;
        min-width: 0;
    }
    .hero-right {
        flex: 2;
        min-width: 0;
        text-align: right;
        display: flex;
        flex-direction: column;
        align-items: flex-end;
        gap: 0.4rem;
    }
    .hero-title {
        font-size: 1.9rem;
        font-weight: 800;
        margin-bottom: 0.15rem;
        letter-spacing: 0.03em;
    }
    .hero-subtitle {
        font-size: 0.98rem;
        color: #9ca3af;
        max-width: 600px;
    }
    .hero-badge {
        display: inline-flex;
        align-items: center;
        gap: 0.4rem;
        padding: 3px 10px;
        border-radius: 999px;
        background: rgba(34,197,94,0.1);
        color: #bbf7d0;
        font-size: 0.75rem;
        font-weight: 600;
        margin-bottom: 0.25rem;
        border: 1px solid rgba(34,197,94,0.35);
    }
    .hero-badge-dot {
        width: 8px;
        height: 8px;
        border-radius: 999px;
        background: #22c55e;
        box-shadow: 0 0 8px rgba(34,197,94,0.9);
    }
    .hero-tagline {
        font-size: 0.9rem;
        color: #e5e7eb;
        font-weight: 500;
    }
    .hero-pill {
        font-size: 0.75rem;
        padding: 4px 10px;
        border-radius: 999px;
        border: 1px solid #4b5563;
        background: rgba(15,23,42,0.7);
        color: #9ca3af;
    }

    @media (max-width: 900px) {
        .hero {
            flex-direction: column;
            align-items: flex-start;
        }
        .hero-right {
            align-items: flex-start;
            text-align: left;
        }
    }

    /* KPI CARDS */
    .kpi-card {
        border-radius: 14px;
        padding: 0.9rem 1rem;
        background: #020617;
        border: 1px solid #1e293b;
        color: #e5e7eb;
        margin-bottom: 0.9rem;
        box-shadow: inset 0 0 18px rgba(0,0,0,0.55);
    }
    .kpi-label {
        font-size: 0.8rem;
        color: #9ca3af;
        margin-bottom: 0.25rem;
        text-transform: uppercase;
        letter-spacing: 0.06em;
    }
    .kpi-main {
        font-size: 1.2rem;
        font-weight: 600;
    }
    .kpi-sub {
        font-size: 0.75rem;
        color: #6b7280;
    }

    /* VALUE BADGES */
    .value-badge {
        display: inline-block;
        padding: 3px 8px;
        border-radius: 6px;
        background: #16a34a;
        color: white;
        font-size: 0.75rem;
        font-weight: 600;
        margin-left: 8px;
    }
    .risk-badge-high {
        display: inline-block;
        padding: 2px 7px;
        border-radius: 999px;
        background: #b91c1c;
        color: white;
        font-size: 0.7rem;
        font-weight: 600;
        margin-left: 6px;
    }
    .risk-badge-medium {
        display: inline-block;
        padding: 2px 7px;
        border-radius: 999px;
        background: #eab308;
        color: #111827;
        font-size: 0.7rem;
        font-weight: 600;
        margin-left: 6px;
    }
    .risk-badge-low {
        display: inline-block;
        padding: 2px 7px;
        border-radius: 999px;
        background: #22c55e;
        color: #052e16;
        font-size: 0.7rem;
        font-weight: 600;
        margin-left: 6px;
    }

    /* MATCH CARD */
    .match-card {
        border-radius: 14px;
        padding: 1rem 1.2rem;
        background: #020617;
        border: 1px solid #1e293b;
        margin-bottom: 1rem;
        color: #e5e7eb;
        box-shadow: 0 0 12px rgba(0,0,0,0.5);
    }
    .match-header {
        font-weight: 700;
        font-size: 1.05rem;
        margin-bottom: 0.25rem;
    }
    .match-sub {
        color: #9ca3af;
        font-size: 0.85rem;
        margin-bottom: 0.6rem;
    }
    .match-row {
        display: flex;
        flex-wrap: wrap;
        justify-content: space-between;
        font-size: 0.85rem;
        gap: 0.4rem;
    }
    .match-col {
        margin-bottom: 0.2rem;
        min-width: 140px;
    }

    /* FILTER BAR */
    .filter-bar {
        padding: 0.6rem 0.8rem;
        border-radius: 12px;
        background: #020617;
        border: 1px solid #1f2937;
        margin-bottom: 0.8rem;
        box-shadow: 0 0 10px rgba(0,0,0,0.4);
    }

    /* TABS */
    button[data-baseweb="tab"] {
        font-size: 0.95rem;
        font-weight: 600 !important;
        color: #e5e7eb !important;
    }

    /* FOOTER */
    .gm-footer {
        margin-top: 1.8rem;
        padding-top: 0.9rem;
        border-top: 1px solid #1f2937;
        font-size: 0.78rem;
        color: #6b7280;
        text-align: center;
    }
    /* ✅ STREAMLIT METRIC "POSVIJETLITI" */
    div[data-testid="stMetric"] {
        background: rgba(30, 41, 59, 0.55) !important;  /* svjetlije od #020617 */
        border: 1px solid rgba(148, 163, 184, 0.35) !important;
        border-radius: 14px !important;
        padding: 14px 14px !important;
        box-shadow: 0 0 14px rgba(0,0,0,0.35) !important;
    }

    div[data-testid="stMetric"] [data-testid="stMetricLabel"] {
        color: #e5e7eb !important;
        font-weight: 700 !important;
        letter-spacing: 0.02em !important;
    }

    div[data-testid="stMetric"] [data-testid="stMetricValue"] {
        color: #ffffff !important;
        font-weight: 800 !important;
        text-shadow: 0 0 10px rgba(255,255,255,0.10) !important;
    }

    div[data-testid="stMetric"] [data-testid="stMetricDelta"] {
        color: #d1d5db !important;
        font-weight: 600 !important;
    }
    /* ✅ GLOBAL: posvijetli sav Markdown tekst (Landing + ostalo) */
[data-testid="stMarkdownContainer"] { 
  color: #e5e7eb !important;
}
[data-testid="stMarkdownContainer"] h1,
[data-testid="stMarkdownContainer"] h2,
[data-testid="stMarkdownContainer"] h3,
[data-testid="stMarkdownContainer"] h4 {
  color: #ffffff !important;
}
[data-testid="stMarkdownContainer"] strong,
[data-testid="stMarkdownContainer"] b {
  color: #ffffff !important;
}
[data-testid="stMarkdownContainer"] li,
[data-testid="stMarkdownContainer"] p {
  color: #e5e7eb !important;
}
[data-testid="stMarkdownContainer"] a {
  color: #93c5fd !important;
}

/* ✅ CAPTION/HELP tekst (često bude preslab) */
[data-testid="stCaptionContainer"] {
  color: #cbd5e1 !important;
}

/* ✅ tvoje custom “kartice” */
.hero, .kpi-card, .match-card, .filter-bar {
  color: #e5e7eb !important;
}
.hero-subtitle, .kpi-sub, .match-sub {
  color: #cbd5e1 !important;
}
/* ✅ SIDEBAR background */
section[data-testid="stSidebar"]{
  background: #000000 !important;
  border-right: 1px solid #1f2937 !important;
}

/* unutarnji wrap (da sve bude crno, ne samo okvir) */
section[data-testid="stSidebar"] > div{
  background: #000000 !important;
}

/* ✅ SIDEBAR text (da ostane čitljivo) */
section[data-testid="stSidebar"] *{
  color: #e5e7eb !important;
}

/* labeli (sliders, selectbox, input) */
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] p{
  color: #e5e7eb !important;
}

/* input/select background (da ne bude bijelo) */
section[data-testid="stSidebar"] input,
section[data-testid="stSidebar"] textarea{
  background: #0b1220 !important;
  color: #e5e7eb !important;
  border: 1px solid #334155 !important;
  border-radius: 10px !important;
}

section[data-testid="stSidebar"] [data-baseweb="select"] > div{
  background: #0b1220 !important;
  color: #e5e7eb !important;
  border: 1px solid #334155 !important;
  border-radius: 10px !important;
}
    </style>
    """, unsafe_allow_html=True)


TEAM_MAPPING_FILE = "team_mapping.xlsx"

def load_team_mapping() -> dict:
    """
    Učitaj team_mapping.xlsx i vrati dict: {fd_name: api_match}.
    """
    path = "team_mapping.xlsx"
    if not os.path.exists(path):
        print("[WARN] team_mapping.xlsx not found – no team mapping will be used.")
        return {}

    df = pd.read_excel(path)

    if "fd_name" not in df.columns or "api_match" not in df.columns:
        print("[WARN] team_mapping.xlsx missing fd_name / api_match columns.")
        return {}

    df["fd_name"] = df["fd_name"].astype(str).str.strip()
    df["api_match"] = df["api_match"].astype(str).str.strip()

    df = df[df["api_match"].notna() & (df["api_match"] != "")]

    mapping = dict(zip(df["fd_name"], df["api_match"]))
    print(f"[OK] Loaded team mapping, rows: {len(mapping)}")

    return mapping

# =========================
# SIMPLE PASSWORD LOGIN
# =========================

def check_password():
    if "authenticated" not in st.session_state:
        st.session_state["authenticated"] = False

    if st.session_state["authenticated"]:
        return True

    st.title("🔐 Login required")

    password = st.text_input("Enter password:", type="password")

    if st.button("Login"):
        if password == "Vice142536":
            st.session_state["authenticated"] = True
            st.success("Login successful!")

            st.query_params["auth"] = "1"
            st.stop()
        else:
            st.error("Wrong password")

    st.stop()


# =========================
# CONFIG
# =========================

DEFAULT_SEASON = "2526"
HISTORICAL_SEASONS = ["2526", "2425", "2324", "2223", "2122"]

ALL_LEAGUES: Dict[str, str] = {
    "Premier League": "E0",
    "Championship": "E1",
    "League One": "E2",
    "League Two": "E3",
    "National League": "EC",
    "Bundesliga": "D1",
    "2. Bundesliga": "D2",
    "Serie A": "I1",
    "Serie B": "I2",
    "La Liga": "SP1",
    "La Liga 2": "SP2",
    "Ligue 1": "F1",
    "Ligue 2": "F2",
    "Eredivisie": "N1",
    "Jupiler Pro League": "B1",
    "Primeira Liga": "P1",
    "Scotland Premier": "SC0",
    "Scotland Championship": "SC1",
    "Scotland League One": "SC2",
    "Scotland League Two": "SC3",
    "Super Lig": "T1",
    "Super League Greece": "G1",
}

BASE_URL = "https://www.football-data.co.uk/mmz4281/{season}/{league_code}.csv"
FIXTURES_URL = "https://www.football-data.co.uk/fixtures.csv"

RAW_FOOTBALL_DIR = os.path.join("data", "raw", "football_data")
os.makedirs(RAW_FOOTBALL_DIR, exist_ok=True)
os.makedirs("models", exist_ok=True)

# API-Football xG cache – očekujemo xg_api_football_*.xlsx unutra
XG_DATA_DIR = os.path.join("data", "api_football")

DC_RHO = 0.13  # Dixon–Coles rho

# === API-FOOTBALL ODDS CONFIG ===

API_FOOTBALL_BASE = "https://v3.football.api-sports.io"

API_LEAGUE_IDS: Dict[str, int] = {
    "Premier League": 39,
    "Championship": 40,
    "League One": 41,
    "League Two": 42,
    "National League": 45,

    "Bundesliga": 78,
    "2. Bundesliga": 79,

    "Serie A": 135,
    "Serie B": 136,

    "La Liga": 140,
    "La Liga 2": 141,

    "Ligue 1": 61,
    "Ligue 2": 62,

    "Eredivisie": 88,
    "Jupiler Pro League": 144,
    "Primeira Liga": 94,

    "Scotland Premier": 179,
    "Scotland Championship": 180,
    "Scotland League One": 181,
    "Scotland League Two": 182,

    "Super Lig": 203,
    "Super League Greece": 197,
}

# Kladionica iz API-Footballa koju želimo koristiti za kvote
PREFERRED_BOOKMAKER = "bet365"  # sve malim slovima, jer u kodu spuštamo na lower()

def season_code_to_year(season_code: str) -> int | None:
    """
    '2526' -> 2025, '2425' -> 2024, ...
    """
    try:
        return 2000 + int(str(season_code)[:2])
    except Exception:
        return None

def get_api_football_key() -> str | None:
    key = None
    # prvo pokušaj iz Streamlit secrets
    try:
        key = st.secrets.get("API_FOOTBALL_KEY", None)
    except Exception:
        key = None
    # onda iz env var
    if not key:
        key = os.getenv("API_FOOTBALL_KEY")
    return key

def safe_float(x):
    try:
        if x is None or x == "":
            return None
        return float(str(x))
    except Exception:
        return None


# =========================
# DOWNLOAD & LOAD HELPERS
# =========================

def download_csv(url: str, dest_path: str) -> None:
    if os.path.exists(dest_path):
        return
    try:
        resp = requests.get(url, timeout=30)
    except Exception as e:
        print(f"[ERR] request failed {url}: {e}")
        return
    if resp.status_code != 200:
        print(f"[ERR] {url} -> {resp.status_code}")
        return
    with open(dest_path, "wb") as f:
        f.write(resp.content)
    print(f"[OK] {url} -> {dest_path}")


def load_all_leagues(season_code: str) -> pd.DataFrame:
    all_dfs = []
    for league_name, league_code in ALL_LEAGUES.items():
        url = BASE_URL.format(season=season_code, league_code=league_code)
        filename = f"{season_code}_{league_code}.csv"
        dest_path = os.path.join(RAW_FOOTBALL_DIR, filename)

        download_csv(url, dest_path)

        if os.path.exists(dest_path):
            try:
                df = pd.read_csv(dest_path, encoding="latin1")
                df["league"] = league_name
                df["league_code"] = league_code
                df["season_code"] = season_code
                all_dfs.append(df)
            except Exception as e:
                print(f"[ERR] loading {dest_path}: {e}")

    if not all_dfs:
        return pd.DataFrame()

    combined = pd.concat(all_dfs, ignore_index=True)
    if "Date" in combined.columns:
        combined["Date"] = pd.to_datetime(combined["Date"], dayfirst=True, errors="coerce")
    else:
        combined["Date"] = pd.NaT

    return combined


def load_all_leagues_multi(seasons: List[str]) -> pd.DataFrame:
    all_dfs = []
    for s in seasons:
        df = load_all_leagues(s)
        if not df.empty:
            all_dfs.append(df)
    if not all_dfs:
        return pd.DataFrame()
    return pd.concat(all_dfs, ignore_index=True)


def load_fixtures_from_web() -> pd.DataFrame:
    try:
        resp = requests.get(FIXTURES_URL, timeout=30)
        if resp.status_code != 200:
            print(f"[WARN] fixtures.csv status {resp.status_code}")
            return pd.DataFrame()

        text = resp.content.decode("latin1", errors="ignore")
        df = pd.read_csv(StringIO(text), sep=None, engine="python")
    except Exception as e:
        print(f"[WARN] load_fixtures_from_web error: {e}")
        return pd.DataFrame()

    cleaned_cols = []
    for c in df.columns:
        c = str(c)
        c = c.replace("\ufeff", "")
        c = c.replace("ï»¿", "")
        c = c.strip()
        cleaned_cols.append(c)
    df.columns = cleaned_cols

    required = ["Div", "Date", "HomeTeam", "AwayTeam"]
    if not all(col in df.columns for col in required):
        print("[WARN] fixtures.csv does not have expected columns:", df.columns.tolist())
        return pd.DataFrame()

    code_to_name = {v: k for k, v in ALL_LEAGUES.items()}
    df["league_code"] = df["Div"].astype(str).str.strip()
    df["league"] = df["league_code"].map(code_to_name)
    df = df[df["league"].notna()].copy()

    df["season_code"] = DEFAULT_SEASON
    df["Date"] = pd.to_datetime(df["Date"], dayfirst=True, errors="coerce")

    if "FTHG" not in df.columns:
        df["FTHG"] = np.nan
    if "FTAG" not in df.columns:
        df["FTAG"] = np.nan

    for col in ["B365H", "B365D", "B365A", "B365>2.5", "B365<2.5"]:
        if col in df.columns:
            df[col] = (
                df[col]
                .astype(str)
                .str.replace(",", ".", regex=False)
                .replace("", np.nan)
                .astype(float)
            )

    for col in ["BTSH", "BTSD"]:
        if col in df.columns:
            df[col] = (
                df[col]
                .astype(str)
                .str.replace(",", ".", regex=False)
                .replace("", np.nan)
                .astype(float)
            )

    return df

XG_DATA_DIR = os.path.join("data", "api_football")

def load_xg_cache() -> pd.DataFrame:
    """
    Učita sve xG fajlove iz data/api_football.
    """
    if not os.path.isdir(XG_DATA_DIR):
        print(f"[WARN] XG_DATA_DIR not found: {XG_DATA_DIR}")
        return pd.DataFrame()

    paths = glob.glob(os.path.join(XG_DATA_DIR, "xg_*.xlsx"))

    if not paths:
        print(f"[WARN] No xg_*.xlsx in {XG_DATA_DIR}")
        return pd.DataFrame()

    dfs = []
    for p in paths:
        try:
            df = pd.read_excel(p)

            rename_map = {}
            for col in df.columns:
                cl = str(col).lower()

                if cl in ["xg_home", "xg_pseudo_home", "xg_home_api"]:
                    rename_map[col] = "xg_home"
                if cl in ["xg_away", "xg_pseudo_away", "xg_away_api"]:
                    rename_map[col] = "xg_away"

                if cl in ["hometeam", "home_team_name", "domacin", "home"]:
                    rename_map[col] = "HomeTeam"
                if cl in ["awayteam", "away_team_name", "gost", "away"]:
                    rename_map[col] = "AwayTeam"

                if cl in ["date", "date_utc"]:
                    rename_map[col] = "Date"

                if cl in ["div", "league_code"]:
                    rename_map[col] = "league_code"

            df = df.rename(columns=rename_map)

            if "Date" in df.columns:
                df["Date"] = pd.to_datetime(df["Date"], errors="coerce")

            if "league_code" not in df.columns:
                fname = os.path.basename(p)  # npr. xg_E0.xlsx
                code = fname.replace("xg_", "").split(".")[0]
                df["league_code"] = code

            if "HomeTeam" in df.columns:
                df["HomeTeam"] = df["HomeTeam"].astype(str).str.strip()
            if "AwayTeam" in df.columns:
                df["AwayTeam"] = df["AwayTeam"].astype(str).str.strip()

            dfs.append(df)
        except Exception as e:
            print(f"[WARN] Cannot load xG file {p}: {e}")

    if not dfs:
        print("[WARN] No valid xG data loaded")
        return pd.DataFrame()

    all_xg = pd.concat(dfs, ignore_index=True)
    print(f"[OK] Loaded xG cache, rows: {all_xg.shape[0]}")
    return all_xg


def merge_xg_into_preds(preds: pd.DataFrame, xg_df: pd.DataFrame) -> pd.DataFrame:
    """
    Spoji xG iz xg_*.xlsx u preds koristeći team_mapping.xlsx.
    """
    if preds.empty or xg_df.empty:
        return preds

    df = preds.copy()
    xg = xg_df.copy()

    team_map = load_team_mapping()

    def map_team(name: str) -> str:
        name_clean = str(name).strip()
        return team_map.get(name_clean, name_clean)

    df["home_api"] = df["home"].apply(map_team).astype(str).str.strip()
    df["away_api"] = df["away"].apply(map_team).astype(str).str.strip()

    if "HomeTeam" in xg.columns:
        xg["HomeTeam"] = xg["HomeTeam"].astype(str).str.strip()
    if "AwayTeam" in xg.columns:
        xg["AwayTeam"] = xg["AwayTeam"].astype(str).str.strip()

    if "league_code" not in df.columns and "league" in df.columns:
        league_to_code = ALL_LEAGUES
        df["league_code"] = df["league"].map(league_to_code)

    if "Date" not in df.columns and "match_date" in df.columns:
        df["Date"] = pd.to_datetime(df["match_date"], errors="coerce")

    if "Date" in xg.columns:
        xg["Date"] = pd.to_datetime(xg["Date"], errors="coerce")

    if "xg_home" not in df.columns:
        df["xg_home"] = np.nan
    if "xg_away" not in df.columns:
        df["xg_away"] = np.nan

    def apply_merge(left: pd.DataFrame,
                    right: pd.DataFrame,
                    left_on: list,
                    right_on: list) -> pd.DataFrame:
        tmp = left.merge(
            right[right_on + ["xg_home", "xg_away"]].drop_duplicates(),
            how="left",
            left_on=left_on,
            right_on=right_on,
            suffixes=("", "_xgtmp"),
        )

        mask = tmp["xg_home"].isna() & tmp["xg_home_xgtmp"].notna()
        tmp.loc[mask, "xg_home"] = tmp.loc[mask, "xg_home_xgtmp"]
        tmp.loc[mask, "xg_away"] = tmp.loc[mask, "xg_away_xgtmp"]

        tmp = tmp.drop(columns=["xg_home_xgtmp", "xg_away_xgtmp"], errors="ignore")
        return tmp

    if all(c in df.columns for c in ["league_code", "Date"]) and \
       all(c in xg.columns for c in ["league_code", "Date", "HomeTeam", "AwayTeam"]):
        df = apply_merge(
            df, xg,
            left_on=["league_code", "Date", "home_api", "away_api"],
            right_on=["league_code", "Date", "HomeTeam", "AwayTeam"],
        )

    if df["xg_home"].isna().any() and \
       ("league_code" in df.columns and "league_code" in xg.columns):
        df = apply_merge(
            df, xg,
            left_on=["league_code", "home_api", "away_api"],
            right_on=["league_code", "HomeTeam", "AwayTeam"],
        )

    if df["xg_home"].isna().any():
        df = apply_merge(
            df, xg,
            left_on=["home_api", "away_api"],
            right_on=["HomeTeam", "AwayTeam"],
        )

    debug_sample = df[
        (df["home"] == "Man City") | (df["away"] == "Man City")
    ][["league", "match_date", "home", "away", "home_api", "away_api", "xg_home", "xg_away"]].head(10)
    print("=== DEBUG Man City mapping ===")
    print(debug_sample)

    df = df.drop(columns=["home_api", "away_api"], errors="ignore")
    return df

import re
import numpy as np
import pandas as pd
import requests
import os

def enrich_with_api_football_odds(df_fixtures: pd.DataFrame, season_code: str) -> pd.DataFrame:
    """
    Bet365-only (ili drugi bookmaker ako postaviš env var), ali STROGO full-time marketi:
      - 1X2: Match Winner
      - OU: Goals Over/Under (FT) za 1.5 i 2.5
      - BTTS: Both Teams Score (FT)

    Env:
      API_FOOTBALL_KEY=...
      API_BOOKMAKER_NAME=Bet365   (opcionalno; default bet365)
      API_BOOKMAKER_ID=8          (opcionalno, ako želiš strogo po ID)
    """
    api_key = get_api_football_key()
    if not api_key:
        print("[API-Football] Nema API ključa – preskačem odds.")
        return df_fixtures

    if df_fixtures is None or df_fixtures.empty:
        print("[API-Football] df_fixtures je prazan – nema što obogatiti kvotama.")
        return df_fixtures

    if "api_fixture_id" not in df_fixtures.columns:
        print("[API-Football] Nema stupca 'api_fixture_id' – ne mogu zvati /odds?fixture=.")
        return df_fixtures

    fixture_ids = sorted({int(x) for x in df_fixtures["api_fixture_id"].dropna().tolist()})
    if not fixture_ids:
        print("[API-Football] Nema nijednog api_fixture_id – preskačem odds.")
        return df_fixtures

    # --- bookmaker filter ---
    bookmaker_name_filter = (os.getenv("API_BOOKMAKER_NAME") or "bet365").strip()
    bookmaker_id_filter = (os.getenv("API_BOOKMAKER_ID") or "").strip()

    def norm(s: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", (s or "").strip().lower())

    bm_name_norm = norm(bookmaker_name_filter)

    headers = {"x-apisports-key": api_key}

    def sfloat(x):
        try:
            if x is None or x == "":
                return np.nan
            return float(str(x).replace(",", "."))
        except Exception:
            return np.nan

    def is_full_time_market(name: str) -> bool:
        n = (name or "").strip().lower()
        bad = [
            "1st half", "first half", "2nd half", "second half",
            "half time", "halftime",
            "home team", "away team",
            "corners", "cards",
        ]

        return not any(b in n for b in bad)

    # strogi match na markete koje želimo
    def is_1x2_market(bet_name: str) -> bool:
        n = (bet_name or "").strip().lower()
        if not is_full_time_market(n):
            return False
        return n in ("match winner", "1x2", "match result", "fulltime result")

    def is_ou_ft_market(bet_name: str) -> bool:
        n = (bet_name or "").strip().lower()
        if not is_full_time_market(n):
            return False
        # samo total goals FT, ne over/under općenito
        return n in ("goals over/under", "goals over/under (total)", "total goals over/under")

    def is_btts_ft_market(bet_name: str) -> bool:
        n = (bet_name or "").strip().lower()
        if not is_full_time_market(n):
            return False
        # samo klasični BTTS FT
        return n in (
            "both teams score",
            "both teams to score",
            "btts",
            "gg/ng",
            "goal/no goal",
            "goal / no goal",
        )

    odds_rows = []

    for fid in fixture_ids:
        markets = {
            "fixture_id": fid,
            "odds_home": np.nan,
            "odds_draw": np.nan,
            "odds_away": np.nan,
            "odds_over15": np.nan,
            "odds_under15": np.nan,
            "odds_over25": np.nan,
            "odds_under25": np.nan,
            "odds_btts_yes": np.nan,
            "odds_btts_no": np.nan,
        }

        try:
            r = requests.get(
                f"{API_FOOTBALL_BASE}/odds",
                headers=headers,
                params={"fixture": fid},
                timeout=20,
            )
        except Exception as e:
            print(f"   ⚠ Greška pri /odds za fixture={fid}: {e}")
            odds_rows.append(markets)
            continue

        if r.status_code != 200:
            print(f"   ⚠ /odds fixture={fid} status={r.status_code}")
            odds_rows.append(markets)
            continue

        try:
            data = r.json()
        except Exception as e:
            print(f"   ⚠ Ne mogu parsirati JSON za fixture={fid}: {e}")
            odds_rows.append(markets)
            continue

        items = data.get("response", []) or []
        if not items:
            odds_rows.append(markets)
            continue

        # ako filtriraš na 1 bookmakera, uzmi baš tog (ne max između marketa)
        def set_if_empty(key, odd):
            if np.isnan(markets[key]) and not np.isnan(odd):
                markets[key] = odd

        for item in items:
            for bm in item.get("bookmakers", []) or []:
                bm_id = str(bm.get("id") or "").strip()
                bm_norm = norm(bm.get("name") or "")

                if bookmaker_id_filter and bm_id != bookmaker_id_filter:
                    continue
                if bm_name_norm and bm_norm != bm_name_norm:
                    continue

                for bet in bm.get("bets", []) or []:
                    bname = bet.get("name") or ""

                    # 1X2
                    if is_1x2_market(bname):
                        for v in bet.get("values") or []:
                            lab = (v.get("value") or "").strip().lower()
                            odd = sfloat(v.get("odd"))
                            if lab in ("home", "1"):
                                set_if_empty("odds_home", odd)
                            elif lab in ("draw", "x"):
                                set_if_empty("odds_draw", odd)
                            elif lab in ("away", "2"):
                                set_if_empty("odds_away", odd)

                    # OU FT (total goals)
                    elif is_ou_ft_market(bname):
                        for v in bet.get("values") or []:
                            lab = (v.get("value") or "").strip().lower()
                            odd = sfloat(v.get("odd"))

                            lab = lab.replace("o ", "over ").replace("u ", "under ")
                            lab = lab.replace("over1.5", "over 1.5").replace("under1.5", "under 1.5")
                            lab = lab.replace("over2.5", "over 2.5").replace("under2.5", "under 2.5")

                            if lab == "over 1.5":
                                set_if_empty("odds_over15", odd)
                            elif lab == "under 1.5":
                                set_if_empty("odds_under15", odd)
                            elif lab == "over 2.5":
                                set_if_empty("odds_over25", odd)
                            elif lab == "under 2.5":
                                set_if_empty("odds_under25", odd)

                    # BTTS FT
                    elif is_btts_ft_market(bname):
                        for v in bet.get("values") or []:
                            lab = (v.get("value") or "").strip().lower()
                            odd = sfloat(v.get("odd"))
                            if lab in ("yes", "y", "gg"):
                                set_if_empty("odds_btts_yes", odd)
                            elif lab in ("no", "n", "ng"):
                                set_if_empty("odds_btts_no", odd)

        odds_rows.append(markets)

    odds_df = pd.DataFrame(odds_rows).drop_duplicates(subset=["fixture_id"], keep="last")

    df_out = df_fixtures.merge(
        odds_df,
        how="left",
        left_on="api_fixture_id",
        right_on="fixture_id",
    )

    # map na tvoje postojeće "best_*"
    df_out["best_home_odds"] = df_out["odds_home"]
    df_out["best_draw_odds"] = df_out["odds_draw"]
    df_out["best_away_odds"] = df_out["odds_away"]

    df_out["best_over15"] = df_out["odds_over15"]
    df_out["best_under15"] = df_out["odds_under15"]
    df_out["best_over25"] = df_out["odds_over25"]
    df_out["best_under25"] = df_out["odds_under25"]

    df_out["best_btts_yes"] = df_out["odds_btts_yes"]
    df_out["best_btts_no"] = df_out["odds_btts_no"]

    n_1x2 = df_out["best_home_odds"].notna().sum()
    n_ou15 = df_out["best_over15"].notna().sum()
    n_ou25 = df_out["best_over25"].notna().sum()
    n_btts = df_out["best_btts_yes"].notna().sum()
    print(f"[DEBUG] API-Football (FT only) odds merged – 1X2: {n_1x2}, OU1.5: {n_ou15}, OU2.5: {n_ou25}, BTTS: {n_btts}")

    return df_out





def load_upcoming_fixtures_api_football(
    season_code: str,
    selected_leagues: List[str],
    days_ahead: int = 5
) -> pd.DataFrame:
    """
    Učita NADOLAZEĆE utakmice iz API-Footballa za odabrane lige
    za period: danas → danas + days_ahead.
    Vraća DF u formatu kompatibilnom s Football-Data:
    league, league_code, season_code, Date, HomeTeam, AwayTeam, FTHG, FTAG, api_fixture_id
    """
    api_key = get_api_football_key()
    if not api_key:
        print("[INFO] API-Football key not configured – no upcoming fixtures from API.")
        return pd.DataFrame()

    season_year = season_code_to_year(season_code)
    if not season_year:
        print(f"[WARN] Cannot infer season year from season_code={season_code}")
        return pd.DataFrame()

    today = datetime.date.today()
    date_from = today
    date_to = today + datetime.timedelta(days=days_ahead)

    from_str = date_from.strftime("%Y-%m-%d")
    to_str = date_to.strftime("%Y-%m-%d")

    headers = {
        "x-apisports-key": api_key,
    }

    rows = []

    for league_name in selected_leagues:
        league_code = ALL_LEAGUES.get(league_name)
        api_league_id = API_LEAGUE_IDS.get(league_name)

        if not league_code or not api_league_id:
            print(f"[WARN] No API-Football league id for league={league_name}")
            continue

        params = {
            "league": api_league_id,
            "season": season_year,
            "from": from_str,
            "to": to_str,
        }

        print(f"➡ [API-Football] Fixtures {league_name} {from_str} → {to_str} ...")

        try:
            r = requests.get(
                f"{API_FOOTBALL_BASE}/fixtures",
                headers=headers,
                params=params,
                timeout=25
            )
        except Exception as e:
            print(f"[ERR] Fixtures request failed for {league_name}: {e}")
            continue

        if r.status_code != 200:
            print(f"[WARN] Fixtures status {r.status_code} for {league_name}")
            continue

        try:
            data = r.json()
            items = data.get("response", [])
        except Exception as e:
            print(f"[ERR] Error parsing fixtures JSON for {league_name}: {e}")
            continue

        print(f"   -> pronađeno fixturea: {len(items)}")

        for it in items:
            fixture = it.get("fixture", {})
            teams = it.get("teams", {})

            home_name = teams.get("home", {}).get("name")
            away_name = teams.get("away", {}).get("name")
            date_str = fixture.get("date")
            fixture_id = fixture.get("id")

            try:
                dt = pd.to_datetime(date_str, errors="coerce")
            except Exception:
                dt = pd.NaT

            rows.append({
                "league": league_name,
                "league_code": league_code,
                "season_code": season_code,
                "Date": dt,
                "HomeTeam": home_name,
                "AwayTeam": away_name,
                "FTHG": np.nan,
                "FTAG": np.nan,
                "api_fixture_id": fixture_id,
            })

    if not rows:
        print("[INFO] No upcoming fixtures from API-Football in given period.")
        return pd.DataFrame()

    df_up = pd.DataFrame(rows)
    return df_up

# =========================
# POISSON + DIXON-COLES
# =========================

def poisson_pmf(k: int, lam: float) -> float:
    return (lam ** k) * exp(-lam) / factorial(k)


def dixon_coles_tau(hg: int, ag: int, lam_home: float, lam_away: float, rho: float) -> float:
    if hg == 0 and ag == 0:
        return 1 - (lam_home + lam_away) * rho
    elif hg == 0 and ag == 1:
        return 1 + lam_home * rho
    elif hg == 1 and ag == 0:
        return 1 + lam_away * rho
    elif hg == 1 and ag == 1:
        return 1 - rho
    else:
        return 1.0


def match_probabilities_dc(
    lam_home: float,
    lam_away: float,
    rho: float = DC_RHO,
    max_goals: int = 10,
) -> Dict[str, float]:
    p_home = p_draw = p_away = 0.0
    total = 0.0

    for hg in range(0, max_goals + 1):
        p_hg = poisson_pmf(hg, lam_home)
        for ag in range(0, max_goals + 1):
            p_ag = poisson_pmf(ag, lam_away)
            base = p_hg * p_ag
            tau = dixon_coles_tau(hg, ag, lam_home, lam_away, rho)
            val = base * tau
            total += val

            if hg > ag:
                p_home += val
            elif hg == ag:
                p_draw += val
            else:
                p_away += val

    if total > 0:
        p_home /= total
        p_draw /= total
        p_away /= total

    return {"p_home": p_home, "p_draw": p_draw, "p_away": p_away}


def goal_market_probs(lam_home: float, lam_away: float, rho: float = DC_RHO, max_goals: int = 10):
    """
    Vraća:
    - p_over15  (zbroj golova >= 2)
    - p_over25  (zbroj golova >= 3)
    - p_btts_yes (oba tima daju gol)
    """
    p_over15 = 0.0
    p_over25 = 0.0
    p_btts = 0.0
    total = 0.0

    for hg in range(0, max_goals + 1):
        p_hg = poisson_pmf(hg, lam_home)
        for ag in range(0, max_goals + 1):
            p_ag = poisson_pmf(ag, lam_away)
            base = p_hg * p_ag
            tau = dixon_coles_tau(hg, ag, lam_home, lam_away, rho)
            p = base * tau
            total += p

            goals = hg + ag
            if goals >= 2:
                p_over15 += p
            if goals >= 3:
                p_over25 += p
            if hg > 0 and ag > 0:
                p_btts += p

    if total > 0:
        p_over15 /= total
        p_over25 /= total
        p_btts /= total

    p_btts_no = 1.0 - p_btts
    return p_over15, p_over25, p_btts, p_btts_no


# =========================
# TEAM STRENGTHS
# =========================

def compute_team_strengths(df: pd.DataFrame) -> pd.DataFrame:
    league_stats = df.groupby("league").agg(
        avg_home_goals=("FTHG", "mean"),
        avg_away_goals=("FTAG", "mean"),
    ).reset_index()

    home_stats = df.groupby(["league", "HomeTeam"]).agg(
        home_goals_for=("FTHG", "sum"),
        home_goals_against=("FTAG", "sum"),
        home_games=("HomeTeam", "count"),
    ).reset_index().rename(columns={"HomeTeam": "team"})

    away_stats = df.groupby(["league", "AwayTeam"]).agg(
        away_goals_for=("FTAG", "sum"),
        away_goals_against=("FTHG", "sum"),
        away_games=("AwayTeam", "count"),
    ).reset_index().rename(columns={"AwayTeam": "team"})

    teams = pd.merge(home_stats, away_stats, on=["league", "team"], how="outer")
    teams = teams.merge(league_stats, on="league", how="left")

    for col in [
        "home_goals_for",
        "home_goals_against",
        "home_games",
        "away_goals_for",
        "away_goals_against",
        "away_games",
    ]:
        teams[col] = teams[col].fillna(0)

    teams["home_games"] = teams["home_games"].replace(0, np.nan)
    teams["away_games"] = teams["away_games"].replace(0, np.nan)

    teams["home_goals_for_avg"] = teams["home_goals_for"] / teams["home_games"]
    teams["home_goals_against_avg"] = teams["home_goals_against"] / teams["home_games"]
    teams["away_goals_for_avg"] = teams["away_goals_for"] / teams["away_games"]
    teams["away_goals_against_avg"] = teams["away_goals_against"] / teams["away_games"]

    teams["home_goals_for_avg"] = teams["home_goals_for_avg"].fillna(teams["avg_home_goals"])
    teams["home_goals_against_avg"] = teams["home_goals_against_avg"].fillna(teams["avg_away_goals"])
    teams["away_goals_for_avg"] = teams["away_goals_for_avg"].fillna(teams["avg_away_goals"])
    teams["away_goals_against_avg"] = teams["away_goals_against_avg"].fillna(teams["avg_home_goals"])

    teams["att_home"] = teams["home_goals_for_avg"] / teams["avg_home_goals"]
    teams["def_home"] = teams["home_goals_against_avg"] / teams["avg_away_goals"]
    teams["att_away"] = teams["away_goals_for_avg"] / teams["avg_away_goals"]
    teams["def_away"] = teams["away_goals_against_avg"] / teams["avg_home_goals"]

    for col in ["att_home", "def_home", "att_away", "def_away"]:
        teams[col] = teams[col].replace([np.inf, -np.inf], np.nan).fillna(1.0)

    return teams


def expected_goals_for_match(
    league_avg_home: float,
    league_avg_away: float,
    home_team_row: pd.Series,
    away_team_row: pd.Series,
) -> Tuple[float, float]:
    lam_home = league_avg_home * home_team_row["att_home"] * away_team_row["def_away"]
    lam_away = league_avg_away * away_team_row["att_away"] * home_team_row["def_home"]
    return max(lam_home, 0.01), max(lam_away, 0.01)


# =========================
# FAIR ODDS + EDGE + KELLY
# =========================

def fair_odds(p: float) -> float:
    if p is None or np.isnan(p) or p <= 0:
        return np.nan
    return 1.0 / p


def compute_edge_and_kelly(p: float, odds: float) -> Tuple[float, float]:
    if p is None or np.isnan(p) or odds is None or np.isnan(odds) or odds <= 1:
        return np.nan, 0.0
    edge = p * odds - 1.0
    if edge <= 0:
        return edge, 0.0
    kelly = edge / (odds - 1.0)
    return edge, kelly


# =========================
# AI FT 1X2 – APPLY
# =========================

def load_ai_1x2_model():
    """
    Učita AI FT 1X2 model iz lokalne datoteke models/ai_1x2_model.pkl.
    """
    model_path = os.path.join("models", "ai_1x2_model.pkl")

    if not os.path.exists(model_path):
        print(f"[ERR] AI 1X2 model not found at {model_path}")
        return None, None

    artifact = joblib.load(model_path)
    return artifact["model"], artifact["feature_cols"]


def apply_ai_model(pred_df: pd.DataFrame) -> pd.DataFrame:
    if pred_df.empty:
        return pred_df

    model, feature_cols = load_ai_1x2_model()
    if model is None or feature_cols is None:
        print("[WARN] AI 1X2 model not available (no .pkl in models/) – ai_p_* = NaN")
        d = pred_df.copy()
        d["ai_p_home"] = np.nan
        d["ai_p_draw"] = np.nan
        d["ai_p_away"] = np.nan
        return d

    df = pred_df.copy()

    if "B365H" not in df.columns and "book_home" in df.columns:
        df["B365H"] = df["book_home"]
    if "B365D" not in df.columns and "book_draw" in df.columns:
        df["B365D"] = df["book_draw"]
    if "B365A" not in df.columns and "book_away" in df.columns:
        df["B365A"] = df["book_away"]

    X = df.reindex(columns=feature_cols, fill_value=0.0)

    prob_matrix = model.predict_proba(X)
    class_to_index = {cls: idx for idx, cls in enumerate(model.classes_)}

    df["ai_p_home"] = prob_matrix[:, class_to_index[0]]
    df["ai_p_draw"] = prob_matrix[:, class_to_index[1]]
    df["ai_p_away"] = prob_matrix[:, class_to_index[2]]

    return df

import numpy as np
import pandas as pd
import os
import streamlit as st
def load_recent_played_fixtures_api_football(
    season_code: str,
    selected_leagues: List[str],
    days_back: int = 7
) -> pd.DataFrame:
    """
    Učita ODIGRANE utakmice iz API-Footballa u periodu: (danas - days_back) → danas.
    Vraća DF kompatibilan s tvojim pipelineom:
    league, league_code, season_code, Date, HomeTeam, AwayTeam, FTHG, FTAG, api_fixture_id
    """
    api_key = get_api_football_key()
    if not api_key:
        print("[INFO] API-Football key not configured – no recent played fixtures from API.")
        return pd.DataFrame()

    season_year = season_code_to_year(season_code)
    if not season_year:
        print(f"[WARN] Cannot infer season year from season_code={season_code}")
        return pd.DataFrame()

    headers = {"x-apisports-key": api_key}

    today = datetime.date.today()
    date_from = today - datetime.timedelta(days=days_back)
    date_to = today

    from_str = date_from.strftime("%Y-%m-%d")
    to_str = date_to.strftime("%Y-%m-%d")

    # mapiraj API team nazive -> FD team nazive (reverse mapping)
    fd_to_api = load_team_mapping()
    api_to_fd = {api_name: fd_name for fd_name, api_name in fd_to_api.items()}

    rows = []
    for league_name in selected_leagues:
        league_code = ALL_LEAGUES.get(league_name)
        api_league_id = API_LEAGUE_IDS.get(league_name)
        if not league_code or not api_league_id:
            continue

        params = {
            "league": api_league_id,
            "season": season_year,
            "from": from_str,
            "to": to_str,
            "status": "FT",   # samo završene
        }

        print(f"➡ [API-Football] Played {league_name} {from_str} → {to_str} (FT) ...")

        try:
            r = requests.get(f"{API_FOOTBALL_BASE}/fixtures", headers=headers, params=params, timeout=25)
            if r.status_code != 200:
                print(f"[WARN] Played fixtures status {r.status_code} for {league_name}")
                continue
            data = r.json()
            items = data.get("response", []) or []
        except Exception as e:
            print(f"[ERR] Played fixtures request failed for {league_name}: {e}")
            continue

        print(f"   -> pronađeno FT: {len(items)}")

        for it in items:
            fixture = it.get("fixture", {})
            teams = it.get("teams", {})
            goals = it.get("goals", {})

            home_api = teams.get("home", {}).get("name")
            away_api = teams.get("away", {}).get("name")

            # map API->FD ako postoji
            home_name = api_to_fd.get(home_api, home_api)
            away_name = api_to_fd.get(away_api, away_api)

            date_str = fixture.get("date")
            fixture_id = fixture.get("id")

            fthg = goals.get("home", None)
            ftag = goals.get("away", None)

            dt = pd.to_datetime(date_str, errors="coerce")

            rows.append({
                "league": league_name,
                "league_code": league_code,
                "season_code": season_code,
                "Date": dt,
                "HomeTeam": home_name,
                "AwayTeam": away_name,
                "FTHG": fthg,
                "FTAG": ftag,
                "api_fixture_id": fixture_id,
            })

    return pd.DataFrame(rows) if rows else pd.DataFrame()
def build_team_form_table(df_played: pd.DataFrame, last_n: int = 5) -> pd.DataFrame:
    """
    Iz odigranih mečeva napravi tablicu forme po timu kroz vrijeme.
    Vraća redove: league, team, match_date, pts, gf, ga, gd,
                  form_pts_lastN, form_gd_lastN
    """
    d = df_played.dropna(subset=["FTHG", "FTAG", "Date"]).copy()
    if d.empty:
        return pd.DataFrame()

    d["Date"] = pd.to_datetime(d["Date"], errors="coerce")
    d = d.dropna(subset=["Date"])

    rows = []

    for _, r in d.iterrows():
        league = r["league"]
        dt = r["Date"]
        h, a = r["HomeTeam"], r["AwayTeam"]
        hg, ag = int(r["FTHG"]), int(r["FTAG"])

        # home row
        h_pts = 3 if hg > ag else (1 if hg == ag else 0)
        rows.append({"league": league, "team": h, "match_date": dt, "pts": h_pts, "gf": hg, "ga": ag, "gd": hg-ag})

        # away row
        a_pts = 3 if ag > hg else (1 if hg == ag else 0)
        rows.append({"league": league, "team": a, "match_date": dt, "pts": a_pts, "gf": ag, "ga": hg, "gd": ag-hg})

    f = pd.DataFrame(rows).sort_values(["league", "team", "match_date"])

    # rolling last N
    f["form_pts_lastN"] = f.groupby(["league","team"])["pts"].transform(lambda s: s.rolling(last_n, min_periods=1).mean())
    f["form_gd_lastN"]  = f.groupby(["league","team"])["gd"].transform(lambda s: s.rolling(last_n, min_periods=1).mean())

    return f
def add_form_strength_to_preds(preds: pd.DataFrame, form_table: pd.DataFrame) -> pd.DataFrame:
    """
    Na preds dodaje:
      home_form_pts, away_form_pts, form_pts_diff
      home_form_gd,  away_form_gd,  form_gd_diff
    Koristi zadnji zapis prije match_date (asof merge po timu).
    """
    if preds is None or preds.empty or form_table is None or form_table.empty:
        return preds

    df = preds.copy()
    df["match_dt"] = pd.to_datetime(df["match_date"], errors="coerce")

    f = form_table.copy()
    f["match_date"] = pd.to_datetime(f["match_date"], errors="coerce")
    f = f.dropna(subset=["match_date"]).sort_values(["league","team","match_date"])

    # merge_asof radi po sortiranom datumu
    def merge_side(side: str):
        out = df.sort_values("match_dt")
        key_team = "home" if side == "home" else "away"

        tmp = pd.merge_asof(
            out.sort_values("match_dt"),
            f.rename(columns={
                "team": key_team,
                "match_date": "hist_dt",
                "form_pts_lastN": f"{side}_form_pts",
                "form_gd_lastN": f"{side}_form_gd",
            }).sort_values("hist_dt"),
            left_on="match_dt",
            right_on="hist_dt",
            by=["league", key_team],
            direction="backward",
            allow_exact_matches=False
        )
        tmp = tmp.drop(columns=["hist_dt"], errors="ignore")
        return tmp

    df = merge_side("home")
    df = merge_side("away")

    df["form_pts_diff"] = df["home_form_pts"] - df["away_form_pts"]
    df["form_gd_diff"]  = df["home_form_gd"]  - df["away_form_gd"]

    df.drop(columns=["match_dt"], errors="ignore", inplace=True)
    return df

# ------------------------------
# PLAYER RATINGS – HELPERI
# ------------------------------

def load_player_stats_for_ratings(path: str = "api_football_player_stats.xlsx") -> pd.DataFrame:
    """
    Učita api_football_player_stats.xlsx i pripremi podatke za računanje team ratinga po utakmicama.
    Fleksibilno traži kolone za rating, team, match_date, minutes.
    Radi i mapiranje liga (E0 -> Premier League) i timova (API naziv -> Football-Data naziv)
    koristeći ALL_LEAGUES i team_mapping.xlsx.
    """
    if not os.path.exists(path):
        print(f"[PLAYER_RAT] Nema fajla {path} – preskačem.")
        return pd.DataFrame()

    try:
        df = pd.read_excel(path)
    except Exception as e:
        print(f"[PLAYER_RAT] Greška pri čitanju {path}: {e}")
        return pd.DataFrame()

    if df.empty:
        print("[PLAYER_RAT] Prazan player stats fajl.")
        return pd.DataFrame()

    # --- rating kolona ---
    rating_candidates = [
        c for c in df.columns
        if c.lower() in ("rating", "player_rating", "score") or "rating" in c.lower()
    ]
    if not rating_candidates:
        print("[PLAYER_RAT] Nema rating kolone – preskačem.")
        return pd.DataFrame()
    rating_col = rating_candidates[0]

    # --- team kolona ---
    team_candidates = [c for c in df.columns if "team" in c.lower()]
    if not team_candidates:
        print("[PLAYER_RAT] Nema team kolone – preskačem.")
        return pd.DataFrame()
    team_col = team_candidates[0]

    # --- match_date / date kolona ---
    date_candidates = [c for c in df.columns if "match_date" in c.lower() or c.lower() == "date"]
    if not date_candidates:
        print("[PLAYER_RAT] Nema match_date/date kolone – preskačem.")
        return pd.DataFrame()
    date_col = date_candidates[0]

    # --- minutes (nije obavezno) ---
    minutes_candidates = [c for c in df.columns if "minute" in c.lower() or "mins" in c.lower()]
    minutes_col = minutes_candidates[0] if minutes_candidates else None

    # --- league kolona (fd_league / league / league_code) ---
    league_candidates = [
        c for c in df.columns
        if c.lower() in ("fd_league", "league", "league_code")
    ]
    league_col = league_candidates[0] if league_candidates else None

    # bazna normalizacija
    df["rating_numeric"] = pd.to_numeric(df[rating_col], errors="coerce")
    df["team_norm"] = df[team_col].astype(str).str.strip()
    df["match_date_norm"] = pd.to_datetime(df[date_col], errors="coerce")

    if minutes_col:
        df["minutes_norm"] = pd.to_numeric(df[minutes_col], errors="coerce")
        df = df[df["minutes_norm"] > 0]
    else:
        df["minutes_norm"] = np.nan

    # --- liga: mapiraj kodove (E0, E1...) na imena ("Premier League"...) ---
    if league_col:
        raw_league = df[league_col].astype(str).str.strip()
        code_to_name = {v: k for k, v in ALL_LEAGUES.items()}  # npr. "E0" -> "Premier League"
        df["league_norm"] = raw_league.map(code_to_name).fillna(raw_league)
    else:
        df["league_norm"] = "UNK"

    # --- team: mapiraj API naziv -> Football-Data naziv preko team_mapping.xlsx ---
    try:
        fd_to_api = load_team_mapping()           # fd_name -> api_match
    except Exception as e:
        print(f"[PLAYER_RAT] Ne mogu učitati team_mapping.xlsx: {e}")
        fd_to_api = {}

    # obrni mapu: API naziv -> FD naziv
    api_to_fd = {api_name: fd_name for fd_name, api_name in fd_to_api.items()}

    df["team_norm"] = df["team_norm"].map(api_to_fd).fillna(df["team_norm"])

    # filtriraj outliere bez datuma/ratinga/tima
    df = df[
        df["match_date_norm"].notna()
        & df["rating_numeric"].notna()
        & df["team_norm"].ne("")
    ].copy()

    if df.empty:
        print("[PLAYER_RAT] Nakon čišćenja nema usable redova.")
        return pd.DataFrame()

    out = df[["league_norm", "team_norm", "match_date_norm", "rating_numeric"]].copy()
    out = out.rename(
        columns={
            "league_norm": "league",
            "team_norm": "team",
            "match_date_norm": "match_date",
            "rating_numeric": "rating",
        }
    )
    print(f"[PLAYER_RAT] Učitano {out.shape[0]} player-rating redova (nakon mapiranja).")
    return out



def build_team_match_rating_history(df_players: pd.DataFrame) -> pd.DataFrame:
    """
    Od player-level ratinga napravi team-match rating:
    prosjek ratinga po timu i utakmici.
    """
    if df_players is None or df_players.empty:
        print("[PLAYER_RAT] Empty df_players – nothing to aggregate.")
        return pd.DataFrame()

    df = df_players.copy()

    required_cols = ["league", "team", "match_date", "rating"]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        print(f"[PLAYER_RAT] Missing columns in df_players: {missing} – preskačem.")
        return pd.DataFrame()

    df = df.dropna(subset=["league", "team", "match_date", "rating"])

    if df.empty:
        print("[PLAYER_RAT] No rows after dropping NA for league/team/date/rating.")
        return pd.DataFrame()

    df = df.groupby(["league", "team", "match_date"], as_index=False).agg(
        team_rating=("rating", "mean"),
        n_players=("rating", "count"),
    )

    df = df.sort_values(["league", "team", "match_date"]).reset_index(drop=True)
    print(f"[PLAYER_RAT] History team-match rating rows: {df.shape[0]}")
    return df


def add_expected_lineup_strength_to_fixtures(
    preds: pd.DataFrame,
    ratings_history: pd.DataFrame,
    last_n_matches: int = 5,
) -> pd.DataFrame:
    """
    Za svaku nadolazeću utakmicu (is_fixture == True) računa:
      - exp_strength_home
      - exp_strength_away
      - exp_strength_diff (home - away)

    Logika:
      - uzme zadnjih N utakmica tog tima prije datuma fixturea
      - izračuna prosjek team_rating (recent_mean)
      - baseline = prosjek team_rating kroz cijelu povijest u fajlu
      - strength_index = recent_mean / baseline   (1.10 = 10% jača postava od prosjeka)
    """
    if preds is None or preds.empty:
        return preds
    if ratings_history is None or ratings_history.empty:
        preds = preds.copy()
        for col in ["exp_strength_home", "exp_strength_away", "exp_strength_diff"]:
            if col not in preds.columns:
                preds[col] = np.nan
        return preds

    df = preds.copy()

    # osiguraj datume
    df["match_date_dt"] = pd.to_datetime(df.get("match_date"), errors="coerce")
    rh = ratings_history.copy()
    rh["match_date_dt"] = pd.to_datetime(rh["match_date"], errors="coerce")
    rh = rh.dropna(subset=["match_date_dt"])

    if rh.empty:
        for col in ["exp_strength_home", "exp_strength_away", "exp_strength_diff"]:
            if col not in df.columns:
                df[col] = np.nan
        df.drop(columns=["match_date_dt"], errors="ignore", inplace=True)
        return df

    # baseline po timu
    baselines = (
        rh.groupby(["league", "team"])["team_rating"]
        .mean()
        .rename("baseline_rating")
        .reset_index()
    )
    rh = rh.merge(baselines, on=["league", "team"], how="left")

    # flag za fixturee
    if "is_fixture" in df.columns:
        df["is_fixture_flag"] = df["is_fixture"].astype(bool)
    else:
        df["is_fixture_flag"] = True

    def compute_strength_for_row(row, side: str):
        league = row.get("league", None)
        team = row.get("home", None) if side == "home" else row.get("away", None)
        match_dt = row.get("match_date_dt", pd.NaT)

        if pd.isna(match_dt) or league is None or pd.isna(league) or team is None or pd.isna(team):
            return np.nan

        hist = rh[
            (rh["league"] == league)
            & (rh["team"] == team)
            & (rh["match_date_dt"] < match_dt)
        ].sort_values("match_date_dt")

        if hist.empty:
            return np.nan

        recent = hist.tail(last_n_matches)
        recent_mean = recent["team_rating"].mean()
        baseline = recent["baseline_rating"].iloc[0]

        if pd.isna(baseline) or baseline <= 0:
            return np.nan

        return recent_mean / baseline

    mask_fix = df["is_fixture_flag"]

    df.loc[mask_fix, "exp_strength_home"] = df[mask_fix].apply(
        lambda r: compute_strength_for_row(r, "home"),
        axis=1,
    )
    df.loc[mask_fix, "exp_strength_away"] = df[mask_fix].apply(
        lambda r: compute_strength_for_row(r, "away"),
        axis=1,
    )

    df["exp_strength_diff"] = df["exp_strength_home"] - df["exp_strength_away"]

    df.drop(columns=["match_date_dt", "is_fixture_flag"], errors="ignore", inplace=True)
    print("[PLAYER_RAT] exp_strength_* dodani u preds.")
    return df


# =========================
# AI GOALS MODELS – TRAIN & APPLY
# =========================

def train_ai_goals_models(df_all: pd.DataFrame) -> None:
    df_played = df_all.dropna(subset=["FTHG", "FTAG"]).copy()
    if df_played.empty:
        print("[ERR] No played matches for goals AI training.")
        return

    teams = compute_team_strengths(df_played)
    league_stats = df_played.groupby("league").agg(
        avg_home_goals=("FTHG", "mean"),
        avg_away_goals=("FTAG", "mean"),
    ).reset_index()

    rows = []
    y_over25 = []
    y_btts = []
    y_total_goals = []

    for _, row in df_played.iterrows():
        league = row["league"]
        home = row["HomeTeam"]
        away = row["AwayTeam"]

        ls = league_stats[league_stats["league"] == league]
        if ls.empty:
            continue
        lg = ls.iloc[0]
        avg_h = lg["avg_home_goals"]
        avg_a = lg["avg_away_goals"]

        ht_rows = teams[(teams["league"] == league) & (teams["team"] == home)]
        at_rows = teams[(teams["league"] == league) & (teams["team"] == away)]
        if ht_rows.empty or at_rows.empty:
            continue

        ht = ht_rows.iloc[0]
        at = at_rows.iloc[0]

        lam_home, lam_away = expected_goals_for_match(avg_h, avg_a, ht, at)
        _, p_over25, p_btts, _ = goal_market_probs(lam_home, lam_away, rho=DC_RHO, max_goals=10)

        fthg = row["FTHG"]
        ftag = row["FTAG"]
        total_goals = fthg + ftag

        label_over25 = 1 if total_goals >= 3 else 0
        label_btts = 1 if (fthg > 0 and ftag > 0) else 0

        feature_row = {
            "lambda_home": lam_home,
            "lambda_away": lam_away,
            "p_over25_poi": p_over25,
            "p_btts_poi": p_btts,
        }

        for col in ["B365H", "B365D", "B365A"]:
            if col in df_played.columns:
                feature_row[col] = row.get(col, np.nan)

        rows.append(feature_row)
        y_over25.append(label_over25)
        y_btts.append(label_btts)
        y_total_goals.append(total_goals)

    X = pd.DataFrame(rows).replace([np.inf, -np.inf], np.nan).fillna(0.0)
    y_over25 = pd.Series(y_over25, name="over25")
    y_btts = pd.Series(y_btts, name="btts")
    y_total_goals = pd.Series(y_total_goals, name="total_goals")

    if X.empty:
        print("[ERR] Goals dataset empty – no training.")
        return

    from sklearn.model_selection import train_test_split
    from sklearn.metrics import accuracy_score

    X_train, X_test, y_train, y_test = train_test_split(
        X, y_over25, test_size=0.25, random_state=42, stratify=y_over25
    )
    over_model = RandomForestClassifier(
        n_estimators=300,
        max_depth=10,
        random_state=42,
        n_jobs=-1,
    )
    print("[INFO] Training RF for Over/Under 2.5...")
    over_model.fit(X_train, y_train)
    y_pred = over_model.predict(X_test)
    acc_over = accuracy_score(y_test, y_pred)
    print(f"[RESULT] Over/Under 2.5 AI accuracy: {acc_over:.3f}")

    X_train2, X_test2, y_train2, y_test2 = train_test_split(
        X, y_btts, test_size=0.25, random_state=42, stratify=y_btts
    )
    btts_model = RandomForestClassifier(
        n_estimators=300,
        max_depth=10,
        random_state=42,
        n_jobs=-1,
    )
    print("[INFO] Training RF for BTTS...")
    btts_model.fit(X_train2, y_train2)
    y_pred2 = btts_model.predict(X_test2)
    acc_btts = accuracy_score(y_test2, y_pred2)
    print(f"[RESULT] BTTS AI accuracy: {acc_btts:.3f}")

    X_train3, X_test3, y_train3, y_test3 = train_test_split(
        X, y_total_goals, test_size=0.25, random_state=42
    )
    goals_model = RandomForestRegressor(
        n_estimators=300,
        max_depth=12,
        random_state=42,
        n_jobs=-1,
    )
    print("[INFO] Training RF regressor for total goals...")
    goals_model.fit(X_train3, y_train3)
    y_pred3 = goals_model.predict(X_test3)
    mae = np.mean(np.abs(y_pred3 - y_test3))
    print(f"[RESULT] Total goals MAE: {mae:.3f}")

    artifact = {
        "feature_cols": list(X.columns),
        "over25_model": over_model,
        "btts_model": btts_model,
        "goals_model": goals_model,
    }
    joblib.dump(artifact, os.path.join("models", "ai_goals_models.pkl"))
    print("[OK] Goals AI models saved to models/ai_goals_models.pkl")


def ensure_ai_goals_models(df_all: pd.DataFrame) -> None:
    model_path = os.path.join("models", "ai_goals_models.pkl")
    if os.path.exists(model_path):
        print("[OK] AI goals models found locally.")
        return

    print("[INFO] ai_goals_models.pkl not found – training locally from historical data...")
    train_ai_goals_models(df_all)


def apply_ai_goals(pred_df: pd.DataFrame) -> pd.DataFrame:
    if pred_df.empty:
        return pred_df

    model_path = os.path.join("models", "ai_goals_models.pkl")
    if not os.path.exists(model_path):
        ensure_ai_goals_models(pred_df.assign(FTHG=np.nan, FTAG=np.nan))

    if not os.path.exists(model_path):
        print("[WARN] AI goals model still not available, using NaN outputs.")
        pred_df = pred_df.copy()
        pred_df["ai_p_over25"] = np.nan
        pred_df["ai_p_under25"] = np.nan
        pred_df["ai_p_btts_yes"] = np.nan
        pred_df["ai_p_btts_no"] = np.nan
        pred_df["ai_total_goals"] = np.nan
        return pred_df

    try:
        artifact = joblib.load(model_path)
    except Exception as e:
        print("[ERR] Failed to load AI goals model:", e)
        pred_df = pred_df.copy()
        pred_df["ai_p_over25"] = np.nan
        pred_df["ai_p_under25"] = np.nan
        pred_df["ai_p_btts_yes"] = np.nan
        pred_df["ai_p_btts_no"] = np.nan
        pred_df["ai_total_goals"] = np.nan
        return pred_df

    feature_cols = artifact["feature_cols"]
    over_model = artifact["over25_model"]
    btts_model = artifact["btts_model"]
    goals_model = artifact["goals_model"]

    df = pred_df.copy()

    if "B365H" not in df.columns and "book_home" in df.columns:
        df["B365H"] = df["book_home"]
    if "B365D" not in df.columns and "book_draw" in df.columns:
        df["B365D"] = df["book_draw"]
    if "B365A" not in df.columns and "book_away" in df.columns:
        df["B365A"] = df["book_away"]

    X = df.reindex(columns=feature_cols, fill_value=0.0)

    prob_over = over_model.predict_proba(X)[:, 1]
    df["ai_p_over25"] = prob_over
    df["ai_p_under25"] = 1.0 - prob_over

    prob_btts = btts_model.predict_proba(X)[:, 1]
    df["ai_p_btts_yes"] = prob_btts
    df["ai_p_btts_no"] = 1.0 - prob_btts

    df["ai_total_goals"] = goals_model.predict(X)

    return df


# =========================
# GENERATE PREDICTIONS
# =========================

def generate_predictions(df_hist: pd.DataFrame, df_current: pd.DataFrame) -> pd.DataFrame:
    preds = []

    df_hist_played = df_hist.dropna(subset=["FTHG", "FTAG"]).copy()
    if df_hist_played.empty or df_current.empty:
        return pd.DataFrame()

    teams = compute_team_strengths(df_hist_played)
    league_stats = df_hist_played.groupby("league").agg(
        avg_home_goals=("FTHG", "mean"),
        avg_away_goals=("FTAG", "mean"),
    ).reset_index()

    for _, row in df_current.iterrows():
        league = row["league"]
        home = row["HomeTeam"]
        away = row["AwayTeam"]

        if pd.isna(home) or pd.isna(away):
            continue

        ls = league_stats[league_stats["league"] == league]
        if ls.empty:
            continue

        lg = ls.iloc[0]
        avg_h = lg["avg_home_goals"]
        avg_a = lg["avg_away_goals"]

        ht_rows = teams[(teams["league"] == league) & (teams["team"] == home)]
        at_rows = teams[(teams["league"] == league) & (teams["team"] == away)]
        if ht_rows.empty or at_rows.empty:
            continue

        ht = ht_rows.iloc[0]
        at = at_rows.iloc[0]

        lam_home, lam_away = expected_goals_for_match(avg_h, avg_a, ht, at)
        probs_1x2 = match_probabilities_dc(lam_home, lam_away, rho=DC_RHO, max_goals=10)

        p_home = probs_1x2["p_home"]
        p_draw = probs_1x2["p_draw"]
        p_away = probs_1x2["p_away"]

        p_over15_poi, p_over25_poi, p_btts_poi, p_btts_no_poi = goal_market_probs(
            lam_home, lam_away, rho=DC_RHO, max_goals=10
        )

        if "Date" in row:
            match_date = row["Date"].date() if pd.notna(row["Date"]) else None
        else:
            match_date = None

        if pd.notna(row.get("FTHG")) and pd.notna(row.get("FTAG")):
            fthg = row["FTHG"]
            ftag = row["FTAG"]
            total_goals_real = fthg + ftag

            if fthg > ftag:
                actual = "H"
            elif fthg == ftag:
                actual = "D"
            else:
                actual = "A"

            actual_over25 = 1 if total_goals_real >= 3 else 0
            actual_btts = 1 if (fthg > 0 and ftag > 0) else 0

            is_fixture = False
        else:
            actual = None
            actual_over25 = None
            actual_btts = None
            is_fixture = True

        # ====== ODDS – COMBO: Football-Data + API-Football MAX ======
        # 1X2
        odd_home = row.get("best_home_odds", np.nan)
        if pd.isna(odd_home):
            odd_home = row.get("B365H", np.nan)

        odd_draw = row.get("best_draw_odds", np.nan)
        if pd.isna(odd_draw):
            odd_draw = row.get("B365D", np.nan)

        odd_away = row.get("best_away_odds", np.nan)
        if pd.isna(odd_away):
            odd_away = row.get("B365A", np.nan)

        # OU 2.5
        book_over25 = row.get("best_over25", np.nan)
        if pd.isna(book_over25):
            book_over25 = row.get("B365>2.5", np.nan)
        book_under25 = row.get("best_under25", np.nan)
        if pd.isna(book_under25):
            book_under25 = row.get("B365<2.5", np.nan)

        # OU 1.5
        book_over15 = row.get("best_over15", np.nan)
        book_under15 = row.get("best_under15", np.nan)

        # BTTS
        book_btts_yes = row.get("best_btts_yes", np.nan)
        if pd.isna(book_btts_yes):
            book_btts_yes = row.get("BTSH", np.nan)

        book_btts_no = row.get("best_btts_no", np.nan)
        if pd.isna(book_btts_no):
            book_btts_no = row.get("BTSD", np.nan)

        # Edge/Kelly FT 1X2
        edge_home, kelly_home = compute_edge_and_kelly(p_home, odd_home)
        edge_draw, kelly_draw = compute_edge_and_kelly(p_draw, odd_draw)
        edge_away, kelly_away = compute_edge_and_kelly(p_away, odd_away)

        if is_fixture:
            model_pick = None
            hit = None
        else:
            if max(p_home, p_draw, p_away) == p_home:
                model_pick = "H"
            elif max(p_home, p_draw, p_away) == p_draw:
                model_pick = "D"
            else:
                model_pick = "A"
            hit = int(model_pick == actual) if actual is not None else None

        preds.append({
            "league": league,
            "match_date": match_date,
            "home": home,
            "away": away,

            "lambda_home": lam_home,
            "lambda_away": lam_away,

            "p_home": p_home,
            "p_draw": p_draw,
            "p_away": p_away,

            "fair_home": fair_odds(p_home),
            "fair_draw": fair_odds(p_draw),
            "fair_away": fair_odds(p_away),

            "book_home": odd_home,
            "book_draw": odd_draw,
            "book_away": odd_away,

            "edge_home": edge_home,
            "edge_draw": edge_draw,
            "edge_away": edge_away,
            "kelly_home": kelly_home,
            "kelly_draw": kelly_draw,
            "kelly_away": kelly_away,

            "p_over15_poi": p_over15_poi,
            "p_over25_poi": p_over25_poi,
            "p_btts_poi": p_btts_poi,
            "p_btts_no_poi": p_btts_no_poi,

            "book_over15": book_over15,
            "book_under15": book_under15,
            "book_over25": book_over25,
            "book_under25": book_under25,
            "book_btts_yes": book_btts_yes,
            "book_btts_no": book_btts_no,

            "actual": actual,
            "actual_over25": actual_over25,
            "actual_btts": actual_btts,
            "is_fixture": is_fixture,
            "model_pick": model_pick,
            "hit": hit,
        })

    return pd.DataFrame(preds)


# =========================
# GOAL VALUE (AI + POISSON) – EDGE & KELLY
# =========================

def add_goal_value_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # === Poisson OU 1.5 ===
    edge_ou15_poi_list = []
    kelly_ou15_poi_list = []
    for _, r in df.iterrows():
        edge, kelly = compute_edge_and_kelly(r.get("p_over15_poi"), r.get("book_over15"))
        edge_ou15_poi_list.append(edge)
        kelly_ou15_poi_list.append(kelly)
    df["edge_ou15_poi"] = edge_ou15_poi_list
    df["kelly_ou15_poi"] = kelly_ou15_poi_list

    # === Poisson OU 2.5 ===
    edge_ou25_poi_list = []
    kelly_ou25_poi_list = []
    for _, r in df.iterrows():
        edge, kelly = compute_edge_and_kelly(r.get("p_over25_poi"), r.get("book_over25"))
        edge_ou25_poi_list.append(edge)
        kelly_ou25_poi_list.append(kelly)
    df["edge_ou25_poi"] = edge_ou25_poi_list
    df["kelly_ou25_poi"] = kelly_ou25_poi_list

    # === Poisson BTTS YES ===
    edge_btts_poi_list = []
    kelly_btts_poi_list = []
    for _, r in df.iterrows():
        edge, kelly = compute_edge_and_kelly(r.get("p_btts_poi"), r.get("book_btts_yes"))
        edge_btts_poi_list.append(edge)
        kelly_btts_poi_list.append(kelly)
    df["edge_btts_poi"] = edge_btts_poi_list
    df["kelly_btts_poi"] = kelly_btts_poi_list

    # === AI OU 2.5 ===
    edges_ou_ai = []
    kelly_ou_ai = []
    for _, r in df.iterrows():
        p = r.get("ai_p_over25")
        o = r.get("book_over25")
        edge, kelly = compute_edge_and_kelly(p, o)
        edges_ou_ai.append(edge)
        kelly_ou_ai.append(kelly)
    df["edge_ou25_ai"] = edges_ou_ai
    df["kelly_ou25_ai"] = kelly_ou_ai

    # === AI BTTS YES ===
    edges_btts_ai = []
    kelly_btts_ai = []
    for _, r in df.iterrows():
        p = r.get("ai_p_btts_yes")
        o = r.get("book_btts_yes")
        edge, kelly = compute_edge_and_kelly(p, o)
        edges_btts_ai.append(edge)
        kelly_btts_ai.append(kelly)
    df["edge_btts_ai"] = edges_btts_ai
    df["kelly_btts_ai"] = kelly_btts_ai

    return df


def add_recommended_and_risk(df: pd.DataFrame) -> pd.DataFrame:
    """
    Dodaje recommended_bet & risk_level.
    """
    df = df.copy()

    kelly_cols = [
        "kelly_home", "kelly_draw", "kelly_away",
        "kelly_ou15_poi", "kelly_ou25_ai", "kelly_btts_ai",
    ]
    for c in kelly_cols:
        if c not in df.columns:
            df[c] = 0.0
        df[c] = df[c].fillna(0.0)

    def row_logic(r):
        k_map = {
            "H": r.get("kelly_home", 0.0),
            "D": r.get("kelly_draw", 0.0),
            "A": r.get("kelly_away", 0.0),
            "OU15": r.get("kelly_ou15_poi", 0.0),
            "OU25": r.get("kelly_ou25_ai", 0.0),
            "BTTS": r.get("kelly_btts_ai", 0.0),
        }

        best_key = max(k_map, key=lambda k: k_map[k])
        best_k = k_map[best_key]

        if best_k <= 0:
            return pd.Series({"recommended_bet": "No bet", "risk_level": "NONE"})

        if best_key == "H":
            rec = "Home win (1)"
        elif best_key == "D":
            rec = "Draw (X)"
        elif best_key == "A":
            rec = "Away win (2)"
        elif best_key == "OU15":
            rec = "Over 1.5 goals"
        elif best_key == "OU25":
            rec = "Over 2.5 goals"
        elif best_key == "BTTS":
            rec = "BTTS YES"
        else:
            rec = "No bet"

        if best_k >= 0.05:
            risk = "HIGH"
        elif best_k >= 0.03:
            risk = "MEDIUM"
        else:
            risk = "LOW"

        return pd.Series({"recommended_bet": rec, "risk_level": risk})

    extra = df.apply(row_logic, axis=1)
    df["recommended_bet"] = extra["recommended_bet"]
    df["risk_level"] = extra["risk_level"]

    return df

def _rec_key_from_text(txt: str) -> str:
    t = (txt or "").strip().lower()
    if t.startswith("home win"):
        return "H"
    if t.startswith("draw"):
        return "D"
    if t.startswith("away win"):
        return "A"
    if "over 1.5" in t:
        return "OU15"
    if "over 2.5" in t:
        return "OU25"
    if "btts" in t:
        return "BTTS"
    return "NONE"


def attach_rec_meta(df: pd.DataFrame) -> pd.DataFrame:
    d = df.copy()
    if "recommended_bet" not in d.columns:
        d["recommended_bet"] = "No bet"

    d["rec_key"] = d["recommended_bet"].apply(_rec_key_from_text)

    # default
    d["rec_p"] = np.nan
    d["rec_odds"] = np.nan
    d["rec_edge"] = np.nan
    d["rec_kelly"] = 0.0

    def set_meta(mask, p_col, o_col, e_col, k_col):
        if p_col in d.columns: d.loc[mask, "rec_p"] = d.loc[mask, p_col]
        if o_col in d.columns: d.loc[mask, "rec_odds"] = d.loc[mask, o_col]
        if e_col in d.columns: d.loc[mask, "rec_edge"] = d.loc[mask, e_col]
        if k_col in d.columns: d.loc[mask, "rec_kelly"] = d.loc[mask, k_col]

    set_meta(d["rec_key"] == "H",    "p_home",        "book_home",    "edge_home",      "kelly_home")
    set_meta(d["rec_key"] == "D",    "p_draw",        "book_draw",    "edge_draw",      "kelly_draw")
    set_meta(d["rec_key"] == "A",    "p_away",        "book_away",    "edge_away",      "kelly_away")
    set_meta(d["rec_key"] == "OU15", "p_over15_poi",  "book_over15",  "edge_ou15_poi",  "kelly_ou15_poi")
    set_meta(d["rec_key"] == "OU25", "ai_p_over25",   "book_over25",  "edge_ou25_ai",   "kelly_ou25_ai")
    set_meta(d["rec_key"] == "BTTS", "ai_p_btts_yes", "book_btts_yes","edge_btts_ai",   "kelly_btts_ai")

    d["rec_kelly"] = pd.to_numeric(d["rec_kelly"], errors="coerce").fillna(0.0)
    return d


def filter_and_risk_two_stage(df: pd.DataFrame) -> pd.DataFrame:
    """
    1) hard_fail: očito krivo / nema inputa / odds out of range
    2) soft_flag: sumnjivi huge-value (prevelik edge/kelly, nesklad modela, čudan total λ)
    -> output: display_bet, display_risk + hard_fail/soft_flag/reasons
    """
    d = attach_rec_meta(df)

    d["hard_fail"] = False
    d["soft_flag"] = False
    d["reasons"] = ""

    def add_reason(mask, msg, hard=False, soft=False):
        if mask.any():
            if hard:
                d.loc[mask, "hard_fail"] = True
            if soft:
                d.loc[mask, "soft_flag"] = True
            d.loc[mask, "reasons"] = d.loc[mask, "reasons"].astype(str)
            d.loc[mask, "reasons"] = np.where(
                d.loc[mask, "reasons"].str.len() > 0,
                d.loc[mask, "reasons"] + " | " + msg,
                msg
            )

    # --- HARD FAILS ---
    m_has_bet = d["rec_key"] != "NONE"

    add_reason(m_has_bet & (d["rec_p"].isna()), "missing rec_p", hard=True)
    add_reason(m_has_bet & (d["rec_odds"].isna()), "missing rec_odds", hard=True)

    add_reason(m_has_bet & (d["rec_odds"] <= 1.01), "odds<=1.01", hard=True)
    add_reason(m_has_bet & (d["rec_odds"] > 25), "odds>25 (bad feed)", hard=True)

    # market-specific sanity
    add_reason((d["rec_key"].isin(["OU15","OU25","BTTS"])) & (d["rec_odds"] > 8), "goals/btts odds>8 (suspicious)", hard=True)

    # --- SOFT FLAGS (huge-value & noise filters) ---
    # Prevelik Kelly/edge često znači mapping/odds problem ili ultra-variance
    add_reason(m_has_bet & (d["rec_edge"].notna()) & (d["rec_edge"] > 0.35), "edge>35% (huge-value)", soft=True)
    add_reason(m_has_bet & (d["rec_kelly"] > 0.12), "kelly>12% (huge-value)", soft=True)

    # Pre-low probability na recommended bet (npr. ‘value’ na 0.07) -> često variance trap
    add_reason(m_has_bet & (d["rec_p"].notna()) & (d["rec_p"] < 0.12), "rec_p<12% (variance trap)", soft=True)

    # Nesklad Poisson vs AI za 1X2
    if all(c in d.columns for c in ["p_home","p_draw","p_away","ai_p_home","ai_p_draw","ai_p_away"]):
        d["poi_pick"] = d[["p_home","p_draw","p_away"]].idxmax(axis=1)
        d["ai_pick2"] = d[["ai_p_home","ai_p_draw","ai_p_away"]].idxmax(axis=1)
        # map to H/D/A
        poi_map = {"p_home":"H","p_draw":"D","p_away":"A"}
        ai_map  = {"ai_p_home":"H","ai_p_draw":"D","ai_p_away":"A"}
        d["poi_pick"] = d["poi_pick"].map(poi_map)
        d["ai_pick2"] = d["ai_pick2"].map(ai_map)

        m_1x2 = d["rec_key"].isin(["H","D","A"])
        add_reason(m_1x2 & (d["poi_pick"] != d["ai_pick2"]) & (d["rec_kelly"] < 0.05),
                   "Poisson≠AI on 1X2 (weak consensus)", soft=True)

    # Goals sanity using λ total (pre-match xG)
    if "xg_pre_total" in d.columns:
        m_ou15 = d["rec_key"] == "OU15"
        add_reason(m_ou15 & (d["xg_pre_total"].notna()) & (d["xg_pre_total"] < 2.10),
                   "OU1.5 but λ_total<2.10", soft=True)

        m_ou25 = d["rec_key"] == "OU25"
        add_reason(m_ou25 & (d["xg_pre_total"].notna()) & (d["xg_pre_total"] < 2.55),
                   "OU2.5 but λ_total<2.55", soft=True)

    # --- DISPLAY RISK (smisleniji) ---
    # base risk from kelly
    def base_risk(k):
        if k >= 0.06: return "HIGH"
        if k >= 0.03: return "MEDIUM"
        if k > 0:     return "LOW"
        return "NONE"

    d["display_risk"] = d["rec_kelly"].apply(base_risk)

    # if soft_flag -> spusti risk za 1 level
    def downgrade(risk: str) -> str:
        if risk == "HIGH": return "MEDIUM"
        if risk == "MEDIUM": return "LOW"
        if risk == "LOW": return "LOW"
        return "NONE"

    d.loc[d["soft_flag"] & (d["display_risk"] != "NONE"), "display_risk"] = d.loc[
        d["soft_flag"] & (d["display_risk"] != "NONE"), "display_risk"
    ].apply(downgrade)

    # hard_fail -> no bet
    d["display_bet"] = d["recommended_bet"]
    d.loc[d["hard_fail"], "display_bet"] = "No bet"
    d.loc[d["hard_fail"], "display_risk"] = "NONE"

    return d


def finalize_picks(df: pd.DataFrame) -> pd.DataFrame:
    # 1) rec + risk_level (tvoje)
    d = add_recommended_and_risk(df)
    # 2) filter + display
    d = filter_and_risk_two_stage(d)
    return d

def betting_settings_sidebar():
    with st.sidebar.expander("🎯 Betting rules (AUTO NO BET / LONGSHOT / Kelly)", expanded=True):
        bank = st.number_input("Bankroll (EUR)", min_value=0.0, value=100.0, step=10.0)
        kelly_mult = st.slider("Kelly multiplier (0.25 = quarter, 0.5 = half)", 0.0, 1.0, 0.50, 0.05)
        max_stake_pct = st.slider("Max stake per bet (% of bank)", 0.5, 25.0, 5.0, 0.5) / 100.0

        st.markdown("**AUTO NO BET thresholds**")
        min_edge = st.slider("Min edge (rec_edge)", 0.0, 0.30, 0.02, 0.01)
        min_kelly = st.slider("Min kelly (rec_kelly)", 0.0, 0.20, 0.01, 0.005)

        st.markdown("**LONGSHOT tagging**")
        long_odds = st.slider("Longshot if odds ≥", 2.0, 15.0, 4.50, 0.25)
        long_pmax = st.slider("...and probability ≤", 0.02, 0.40, 0.18, 0.01)

    return {
        "bank": float(bank),
        "kelly_mult": float(kelly_mult),
        "max_stake_pct": float(max_stake_pct),
        "min_edge": float(min_edge),
        "min_kelly": float(min_kelly),
        "long_odds": float(long_odds),
        "long_pmax": float(long_pmax),
    }


def apply_betting_rules(df: pd.DataFrame, settings: dict) -> pd.DataFrame:
    """
    Dodaje:
      - auto_action: NO BET / BET / LONGSHOT
      - stake_eur: Kelly stake u EUR (bank * kelly * multiplier, capped)
    Očekuje da df već ima finalize_picks() output. Ako nema, pozove ga.
    """
    d = df.copy()

    # ako nema output od finalize_picks, pokušaj ga napraviti
    needed = {"display_bet", "hard_fail", "rec_edge", "rec_kelly", "rec_odds", "rec_p"}
    if not needed.issubset(set(d.columns)):
        d = finalize_picks(d)

    bank = settings["bank"]
    kelly_mult = settings["kelly_mult"]
    max_stake = bank * settings["max_stake_pct"]
    min_edge = settings["min_edge"]
    min_kelly = settings["min_kelly"]
    long_odds = settings["long_odds"]
    long_pmax = settings["long_pmax"]

    # default
    d["auto_action"] = "NO BET"

    # kandidat za bet = nije hard_fail + ima display_bet
    can_bet = (~d.get("hard_fail", False)) & (d.get("display_bet", "No bet") != "No bet")

    # AUTO NO BET (pragovi)
    pass_thresholds = (
        can_bet
        & d["rec_edge"].notna() & (d["rec_edge"] >= min_edge)
        & d["rec_kelly"].notna() & (d["rec_kelly"] >= min_kelly)
    )

    d.loc[pass_thresholds, "auto_action"] = "BET"

    # LONGSHOT tag (samo ako je već BET)
    is_longshot = (
        (d["auto_action"] == "BET")
        & d["rec_odds"].notna() & (d["rec_odds"] >= long_odds)
        & d["rec_p"].notna() & (d["rec_p"] <= long_pmax)
    )
    d.loc[is_longshot, "auto_action"] = "LONGSHOT"

    # Kelly stake (EUR)
    d["stake_eur"] = 0.0
    raw_stake = bank * d["rec_kelly"].fillna(0.0) * kelly_mult
    raw_stake = raw_stake.clip(lower=0.0, upper=max_stake)

    d.loc[d["auto_action"].isin(["BET", "LONGSHOT"]), "stake_eur"] = raw_stake

    return d


# =========================
# EXCEL PRO – FIXTURES
# =========================

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

def build_pro_fixtures_excel(fixtures_dashboard: pd.DataFrame, season: str) -> BytesIO:
    """
    PRO Excel v2.
    """

    buffer_fix = BytesIO()

    fixtures_dashboard = finalize_picks(fixtures_dashboard)

    sort_cols = [c for c in ["match_date", "league"] if c in fixtures_dashboard.columns]
    if sort_cols:
        fixtures_dashboard = fixtures_dashboard.sort_values(sort_cols)

    with pd.ExcelWriter(buffer_fix, engine="openpyxl") as writer:
        fixtures_dashboard.to_excel(
            writer,
            sheet_name="Fixtures_PRO",
            index=False,
            startrow=3
        )

        wb = writer.book
        ws = writer.sheets["Fixtures_PRO"]

        max_col = fixtures_dashboard.shape[1]
        last_col_letter = get_column_letter(max_col)
        col_names = list(fixtures_dashboard.columns)

        title_cell = ws["A1"]
        ws.merge_cells(f"A1:{last_col_letter}1")
        title_cell.value = f"GOALMIND PRO – Fixtures (Poisson + AI + Kelly) {season}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        info_cell = ws["A2"]
        ws.merge_cells(f"A2:{last_col_letter}2")
        info_cell.value = (
            "Expected goals (λ), Poisson & AI FT 1X2, AI goals (OU / BTTS), fair odds, edge & Kelly, "
            "recommended bet (1X2 / OU / BTTS) + risk level."
        )
        info_cell.font = Font(size=10, italic=True)
        info_cell.alignment = Alignment(horizontal="center", vertical="center")

        section_row = 3
        header_row = 4

        def cols_for(names):
            return [col_names.index(n) + 1 for n in names if n in col_names]

        sections = [
            ("Match info", cols_for(["league", "match_date", "home", "away"])),

            ("Signals", cols_for(["display_bet", "display_risk", "reasons"])),

            ("Poisson FT 1X2", cols_for(["lambda_home", "lambda_away", "p_home", "p_draw", "p_away"])),

            ("AI FT 1X2", cols_for(["ai_p_home", "ai_p_draw", "ai_p_away"])),

            ("Poisson goals", cols_for(["p_over15_poi", "p_over25_poi", "p_btts_poi", "p_btts_no_poi"])),

            ("AI goals", cols_for(["ai_p_over25", "ai_p_btts_yes", "ai_total_goals"])),

            ("Odds", cols_for([
                "book_home", "book_draw", "book_away",
                "book_over15", "book_over25",
                "book_btts_yes", "book_btts_no"
            ])),

            ("Edge & Kelly", cols_for([
                "edge_home", "edge_draw", "edge_away",
                "kelly_home", "kelly_draw", "kelly_away",
                "edge_ou15_poi", "kelly_ou15_poi",
                "edge_ou25_poi", "kelly_ou25_poi",
                "edge_ou25_ai", "kelly_ou25_ai",
                "edge_btts_poi", "kelly_btts_poi",
                "edge_btts_ai", "kelly_btts_ai",
            ])),
        ]

        section_fill = PatternFill("solid", fgColor="E5E7EB")
        section_font = Font(bold=True)

        for label, cols_idx in sections:
            if not cols_idx:
                continue
            start_col = min(cols_idx)
            end_col = max(cols_idx)
            start_letter = get_column_letter(start_col)
            end_letter = get_column_letter(end_col)
            ws.merge_cells(f"{start_letter}{section_row}:{end_letter}{section_row}")
            cell = ws[f"{start_letter}{section_row}"]
            cell.value = label
            cell.font = section_font
            cell.fill = section_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="CCCCCC")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            col_letter = get_column_letter(col_idx)
            header_name = col_names[col_idx - 1].lower()

            if col_letter in ["A", "B"]:
                ws.column_dimensions[col_letter].width = 14
            elif col_letter in ["C", "D"]:
                ws.column_dimensions[col_letter].width = 22
            elif header_name in ["recommended_bet", "risk_level"]:
                ws.column_dimensions[col_letter].width = 20
            else:
                ws.column_dimensions[col_letter].width = 12

        last_row = ws.max_row

        for row_idx in range(header_row + 1, last_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = ws["A5"]
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_row}"

        edge_cols = [i + 1 for i, c in enumerate(col_names) if c.startswith("edge_")]
        kelly_cols = [i + 1 for i, c in enumerate(col_names) if c.startswith("kelly_")]

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        for col_idx in edge_cols:
            col_letter = get_column_letter(col_idx)
            rng = f"{col_letter}{header_row + 1}:{col_letter}{last_row}"
            ws.conditional_formatting.add(
                rng,
                CellIsRule(operator='greaterThanOrEqual', formula=['0.05'], fill=green_fill)
            )

        for col_idx in kelly_cols:
            col_letter = get_column_letter(col_idx)
            rng = f"{col_letter}{header_row + 1}:{col_letter}{last_row}"
            ws.conditional_formatting.add(
                rng,
                CellIsRule(operator='greaterThanOrEqual', formula=['0.03'], fill=green_fill)
            )
            ws.conditional_formatting.add(
                rng,
                CellIsRule(operator='between', formula=['0.015', '0.03'], fill=yellow_fill)
            )

        risk_col_idx = col_names.index("risk_level") + 1 if "risk_level" in col_names else None
        rec_col_idx = col_names.index("recommended_bet") + 1 if "recommended_bet" in col_names else None

        fill_high = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
        fill_med = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        fill_low = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_none = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        for row_idx in range(header_row + 1, last_row + 1):
            if risk_col_idx:
                cell = ws.cell(row=row_idx, column=risk_col_idx)
                val = str(cell.value).upper() if cell.value is not None else ""
                if val == "HIGH":
                    cell.fill = fill_high
                elif val == "MEDIUM":
                    cell.fill = fill_med
                elif val == "LOW":
                    cell.fill = fill_low
                elif val == "NONE":
                    cell.fill = fill_none

            if rec_col_idx:
                rec_cell = ws.cell(row=row_idx, column=rec_col_idx)
                rec_cell.font = Font(bold=True)

        best_bets = fixtures_dashboard[
            (fixtures_dashboard.get("display_bet", fixtures_dashboard["recommended_bet"]) != "No bet") &
            (fixtures_dashboard.get("display_risk", fixtures_dashboard["risk_level"]).isin(["HIGH", "MEDIUM", "LOW"])) &
            (~fixtures_dashboard.get("hard_fail", False))
            ].copy()

        if not best_bets.empty:
            risk_order = {"HIGH": 3, "MEDIUM": 2, "LOW": 1, "NONE": 0}
            best_bets["risk_order"] = best_bets["risk_level"].map(risk_order)

            kelly_cols_all = [c for c in best_bets.columns if c.startswith("kelly_")]
            if kelly_cols_all:
                best_bets["max_kelly_any"] = best_bets[kelly_cols_all].max(axis=1)
            else:
                best_bets["max_kelly_any"] = 0.0

            best_bets = best_bets.sort_values(
                by=["risk_order", "max_kelly_any"], ascending=[False, False]
            )

            cols_best = [
                "league", "match_date", "home", "away",
                "recommended_bet", "risk_level",
                "p_home", "p_draw", "p_away",
                "ai_p_home", "ai_p_draw", "ai_p_away",
                "p_over15_poi", "p_over25_poi",
                "ai_p_over25",
                "p_btts_poi", "ai_p_btts_yes",
                "book_home", "book_draw", "book_away",
                "book_over15", "book_over25",
                "book_btts_yes", "book_btts_no",
                "edge_home", "edge_draw", "edge_away",
                "edge_ou15_poi", "edge_ou25_poi", "edge_ou25_ai",
                "edge_btts_poi", "edge_btts_ai",
                "kelly_home", "kelly_draw", "kelly_away",
                "kelly_ou15_poi", "kelly_ou25_poi", "kelly_ou25_ai",
                "kelly_btts_poi", "kelly_btts_ai",
            ]
            cols_best = [c for c in cols_best if c in best_bets.columns]

            best_bets[cols_best].to_excel(
                writer,
                sheet_name="Best_Bets",
                index=False,
                startrow=0
            )

            ws2 = writer.sheets["Best_Bets"]
            max_col2 = len(cols_best)
            last_col_letter2 = get_column_letter(max_col2)

            header_font2 = Font(bold=True)
            header_fill2 = PatternFill("solid", fgColor="D9E1F2")

            for col_idx in range(1, max_col2 + 1):
                cell = ws2.cell(row=1, column=col_idx)
                cell.font = header_font2
                cell.fill = header_fill2
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

                col_letter = get_column_letter(col_idx)
                name = cols_best[col_idx - 1].lower()
                if col_letter in ["A", "B"]:
                    ws2.column_dimensions[col_letter].width = 14
                elif col_letter in ["C", "D"]:
                    ws2.column_dimensions[col_letter].width = 22
                else:
                    ws2.column_dimensions[col_letter].width = 14

            last_row2 = ws2.max_row
            for row_idx in range(2, last_row2 + 1):
                for col_idx in range(1, max_col2 + 1):
                    cell = ws2.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal="center", vertical="center")

            risk2_idx = cols_best.index("risk_level") + 1 if "risk_level" in cols_best else None
            rec2_idx = cols_best.index("recommended_bet") + 1 if "recommended_bet" in cols_best else None

            for row_idx in range(2, last_row2 + 1):
                if risk2_idx:
                    cell = ws2.cell(row=row_idx, column=risk2_idx)
                    val = str(cell.value).upper() if cell.value is not None else ""
                    if val == "HIGH":
                        cell.fill = fill_high
                    elif val == "MEDIUM":
                        cell.fill = fill_med
                    elif val == "LOW":
                        cell.fill = fill_low
                    elif val == "NONE":
                        cell.fill = fill_none

                if rec2_idx:
                    rec_cell = ws2.cell(row=row_idx, column=rec2_idx)
                    rec_cell.font = Font(bold=True)

            ws2.auto_filter.ref = f"A1:{last_col_letter2}{last_row2}"
            ws2.freeze_panes = ws2["A2"]

        ws3 = wb.create_sheet("Info")

        ws3["A1"] = "GOALMIND PRO – Fixtures Excel v2"
        ws3["A1"].font = Font(bold=True, size=14)

        ws3["A3"] = "How to use this file:"
        ws3["A3"].font = Font(bold=True)

        ws3["A4"] = "- Sheet 'Fixtures_PRO': full fixtures table with Poisson, AI, odds, edge & Kelly and recommended bets."
        ws3["A5"] = "- Sheet 'Best_Bets': filtered shortlist of highest quality value bets sorted by risk level and Kelly."
        ws3["A6"] = "- Use filters on header row to filter by league, date, risk_level, market type, etc."
        ws3["A7"] = "- Green cells in edge/Kelly columns = strong value. Yellow = medium value."

        ws3["A9"] = "Risk level legend:"
        ws3["A9"].font = Font(bold=True)
        ws3["A10"] = "HIGH  = agresivni value (veći Kelly, veći swing)."
        ws3["A11"] = "MEDIUM = balansirano (dobar value, ali ne ekstremno)."
        ws3["A12"] = "LOW   = manji edge, više za lean / fun stakes."
        ws3["A13"] = "NONE  = nema dovoljno edge-a – preskoči."

        ws3["A15"] = "Key columns:"
        ws3["A15"].font = Font(bold=True)

        explanations = [
            ("league", "League name (e.g., Championship, League One, Ligue 2...)."),
            ("match_date", "Match date (YYYY-MM-DD)."),
            ("home / away", "Home and away team names."),
            ("lambda_home / lambda_away", "Expected goals (Poisson λ) for each team."),
            ("xg_pre_home / xg_pre_away / xg_pre_total", "Pre-match xG-style expectation based on Poisson model."),
            ("p_home / p_draw / p_away", "Poisson+Dixon-Coles probability for FT 1X2."),
            ("ai_p_home / ai_p_draw / ai_p_away", "AI model probability for FT 1X2 (trained on multi-season data)."),
            ("p_over15_poi / p_over25_poi", "Poisson probability for Over 1.5 / Over 2.5."),
            ("p_btts_poi / p_btts_no_poi", "Poisson probability for BTTS YES / NO."),
            ("ai_p_over25 / ai_p_btts_yes / ai_total_goals", "AI probabilities and expected total goals."),
            ("book_*", "Market odds (FT 1X2, OU 1.5, OU 2.5, BTTS Yes/No)."),
            ("edge_*", "Value indicator: edge = p * odds - 1. If > 0, model sees value."),
            ("kelly_*", "Kelly fraction for stake sizing. Typical use: stake = bank * Kelly * safety_factor."),
            ("recommended_bet", "Main suggestion for that match (1X2 / OU / BTTS YES / No bet)."),
            ("risk_level", "Risk profile of the recommended bet: LOW / MEDIUM / HIGH / NONE."),
        ]

        start_row = 17
        ws3["A16"] = "Column"
        ws3["B16"] = "Description"
        ws3["A16"].font = Font(bold=True)
        ws3["B16"].font = Font(bold=True)

        for i, (col_name, desc) in enumerate(explanations):
            r = start_row + i
            ws3[f"A{r}"] = col_name
            ws3[f"B{r}"] = desc

        ws3.column_dimensions["A"].width = 26
        ws3.column_dimensions["B"].width = 100

    buffer_fix.seek(0)
    return buffer_fix


# =========================
# ROI HELPER
# =========================

def compute_roi_binary(
    df: pd.DataFrame,
    prob_col: str,
    odds_col: str,
    actual_col: str,
    edge_threshold: float = 0.0
):
    d = df.copy()
    d = d[
        d[prob_col].notna()
        & d[odds_col].notna()
        & d[actual_col].notna()
    ]

    if d.empty:
        return np.nan, 0

    d["edge_tmp"] = d[prob_col] * d[odds_col] - 1.0
    d = d[d["edge_tmp"] > edge_threshold]

    if d.empty:
        return np.nan, 0

    d["profit"] = np.where(
        d[actual_col] == 1,
        d[odds_col] - 1.0,
        -1.0
    )

    total_profit = d["profit"].sum()
    n_bets = d.shape[0]
    roi = total_profit / n_bets

    return roi, n_bets


XG_DATA_DIR = os.path.join("data", "api_football")


# =========================
# LANDING PAGE
# =========================

def render_landing_page():
    st.markdown("""
    <div style='text-align:center; padding: 15px 0;'>
        <h1 style='margin-bottom:0;'>⚡ GOALMIND PRO</h1>
        <p style='font-size:17px; color:#6b7280;'>Poisson • Dixon–Coles • AI • xG • Kelly • Value Bets</p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("---")

    st.subheader("📘 How to Use GOALMIND PRO")

    st.markdown("""
    **GOALMIND PRO is built for bettors, analysts, and football data users who want fast, accurate and automated insights.**

    - Poisson + Dixon–Coles = baseline probabilities  
    - AI models = pattern recognition on multi-season data  
    - xG (from API-Football) = real underlying chance quality  
    - Kelly + edge = staking + value picking

    Use the tabs above to move between:
    - Overview (global KPIs)
    - xG analysis (real chance creation vs goals)
    - FT 1X2 details
    - Goals OU/BTTS details
    - Fixtures & value bets
    - Excel exports
    """)

    st.markdown("---")
    st.markdown(
        "<p style='text-align:center; color:#9ca3af;'>© 2025 GOALMIND PRO – Football Predictions Engine</p>",
        unsafe_allow_html=True
    )
def add_last_round_strength_to_preds(preds: pd.DataFrame, df_played_all_for_strength: pd.DataFrame, last_n: int = 1) -> pd.DataFrame:
    # 1) od raw odigranih napravi form tablicu (last_n=1 = samo zadnja utakmica)
    form_table = build_team_form_table(df_played_all_for_strength, last_n=last_n)

    # 2) form tablicu spoji u preds
    return add_form_strength_to_preds(preds, form_table)


# =========================
# STREAMLIT APP
# =========================

def main():
    st.set_page_config(page_title="GOALMIND PRO – Poisson + AI + xG", layout="wide")

    inject_pro_css()

    # SIDEBAR – prvo da imamo season za hero
    st.sidebar.title("⚙️ Model settings")

    season = st.sidebar.text_input("Season (football-data code)", value=DEFAULT_SEASON)
    selected_leagues = st.sidebar.multiselect(
        "Leagues",
        options=list(ALL_LEAGUES.keys()),
        default=list(ALL_LEAGUES.keys()),
    )

    min_edge = st.sidebar.slider("Min edge (FT 1X2)", 0.0, 0.2, 0.02, 0.005)
    min_kelly = st.sidebar.slider("Min Kelly (FT 1X2)", 0.0, 0.1, 0.01, 0.005)
    bet_settings = betting_settings_sidebar()

    st.sidebar.markdown("---")
    api_key_input = st.sidebar.text_input("API-Football key (optional)", type="password", help="Ako upišeš ovdje, koristi se samo za ovaj run.")
    if api_key_input:
        os.environ["API_FOOTBALL_KEY"] = api_key_input

    st.sidebar.write("Data source: Football-Data.co.uk (history + fixtures)")
    st.sidebar.write("xG & odds source: API-Football (post-match stats, odds)")

    # HERO HEADER – sada season stvarno radi
    st.markdown(
        f"""
        <div class="hero">
            <div class="hero-left">
                <div class="hero-badge">
                    <div class="hero-badge-dot"></div>
                    LIVE MODEL • MULTI-LEAGUE • PRO
                </div>
                <div class="hero-title">⚡ GOALMIND PRO</div>
                <div class="hero-subtitle">
                    Poisson + Dixon–Coles + AI + xG + Kelly • FT 1X2 • OU 1.5 & 2.5 • BTTS • Value bets for serious bettors.
                </div>
            </div>
            <div class="hero-right">
                <div class="hero-tagline">Season {season} • Football-Data + API-Football</div>
                <div class="hero-pill">Made by Vice Maslov • BETA</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # 1) History
    with st.spinner("Loading historical data (multi-season)..."):
        df_hist = load_all_leagues_multi(HISTORICAL_SEASONS)
        if df_hist.empty:
            st.error("No historical data found – check your files / connection.")
            return
        df_hist = df_hist[df_hist["league"].isin(selected_leagues)]

    # 2) Train goals AI if missing
    ensure_ai_goals_models(df_hist)

    with st.spinner("Loading current season and fixtures..."):
        # 1) Rezultati tekuće sezone iz Football-Data (povijest)
        df_curr = load_all_leagues(season)

        if not df_curr.empty:
            df_curr = df_curr[df_curr["league"].isin(selected_leagues)]

        # 2) NADOLAZEĆE utakmice iz API-Football (umjesto fixtures.csv)
        upcoming_api = load_upcoming_fixtures_api_football(
            season_code=season,
            selected_leagues=selected_leagues,
            days_ahead=5,  # npr. 5 dana unaprijed – po želji promijeni
        )

        if not upcoming_api.empty:
            df_curr = pd.concat([df_curr, upcoming_api], ignore_index=True)

        if df_curr.empty:
            df_played_current = pd.DataFrame()
            df_fixtures_current = pd.DataFrame()
        else:
            df_played_current = df_curr.dropna(subset=["FTHG", "FTAG"]).copy()
            df_fixtures_current = df_curr[
                df_curr["FTHG"].isna()
                & df_curr["FTAG"].isna()
                & df_curr["HomeTeam"].notna()
                & df_curr["AwayTeam"].notna()
            ].copy()
        # =========================
        # (NOVO) PLAYED DATA ZA "KOLO PRIJE" / FORMA
        # =========================
        df_hist_played = df_hist.dropna(subset=["FTHG", "FTAG"]).copy()

        # Spoji historiju + current played (da imaš što više mečeva za tim)
        df_played_all_for_strength = pd.concat(
            [df_hist_played, df_played_current],
            ignore_index=True
        )

        # osiguraj Date datetime
        df_played_all_for_strength["Date"] = pd.to_datetime(
            df_played_all_for_strength["Date"], dayfirst=True, errors="coerce"
        )

        # 3) API-FOOTBALL ODDS za te fixtures (kao i prije)
        if not df_fixtures_current.empty:
            df_fixtures_current = enrich_with_api_football_odds(df_fixtures_current, season_code=season)

        raw_fixtures_count = df_fixtures_current.shape[0]

        preds_played = generate_predictions(df_hist, df_played_current) if not df_played_current.empty else pd.DataFrame()
        preds_fixtures = generate_predictions(df_hist, df_fixtures_current) if not df_fixtures_current.empty else pd.DataFrame()

        if preds_played.empty and preds_fixtures.empty:
            st.warning("No predictions to display.")
            return
        elif preds_played.empty:
            preds = preds_fixtures.copy()
        elif preds_fixtures.empty:
            preds = preds_played.copy()
        else:
            preds = pd.concat([preds_played, preds_fixtures], ignore_index=True)
    # =========================
    # (NOVO) DODAJ "KOLO PRIJE" / FORMA U PREDS
    # =========================
    preds = add_last_round_strength_to_preds(
        preds,
        df_played_all_for_strength,
        last_n=1  # 1 = samo kolo prije; stavi 5 za formu
    )

    # 4) AI FT 1X2 + AI goals + Kelly for goals
    preds = apply_ai_model(preds)
    preds = apply_ai_goals(preds)
    preds = add_goal_value_columns(preds)
    # === 1X2 RECO (Poisson vs AI blend) ===
    # Poisson 1X2 u tvom preds-u su: p_home, p_draw, p_away
    preds["p_poi_h"] = preds["p_home"]
    preds["p_poi_d"] = preds["p_draw"]
    preds["p_poi_a"] = preds["p_away"]

    # AI 1X2 u tvom preds-u su: ai_p_home, ai_p_draw, ai_p_away
    preds["p_ai_h"] = preds["ai_p_home"]
    preds["p_ai_d"] = preds["ai_p_draw"]
    preds["p_ai_a"] = preds["ai_p_away"]

    # Odds u tvom preds-u su: book_home, book_draw, book_away
    preds["odds_h"] = preds["book_home"]
    preds["odds_d"] = preds["book_draw"]
    preds["odds_a"] = preds["book_away"]

    # Preporuku radi smisleno samo za fixturee (nadolazeće)
    mask_fix = preds["is_fixture"] == True

    tmp = preds.loc[mask_fix].apply(lambda r: pick_1x2_reco(r, w_ai=0.55), axis=1)

    preds.loc[mask_fix, "reco_1x2"] = tmp.apply(lambda x: x[0])
    preds.loc[mask_fix, "conf_1x2"] = tmp.apply(lambda x: x[1])
    preds.loc[mask_fix, "value_rel_1x2"] = tmp.apply(lambda x: x[2])
    preds.loc[mask_fix, "edge_abs_1x2"] = tmp.apply(lambda x: x[3])
    preds.loc[mask_fix, "kelly_1x2"] = tmp.apply(lambda x: x[4])

    TEAM_MAP = load_team_mapping()

    # 5) xG – učitaj cache i spoji u preds
    xg_cache = load_xg_cache()
    preds = merge_xg_into_preds(preds, xg_cache)

    preds["xg_pre_home"] = preds["lambda_home"]
    preds["xg_pre_away"] = preds["lambda_away"]
    preds["xg_pre_total"] = preds["xg_pre_home"] + preds["xg_pre_away"]

    # 5b) Expected lineup strength na temelju player ratinga (zadnjih X utakmica)
    try:
        df_players = load_player_stats_for_ratings("api_football_player_stats.xlsx")
        ratings_history = build_team_match_rating_history(df_players)

        preds = add_expected_lineup_strength_to_fixtures(
            preds,
            ratings_history,
            last_n_matches=5,  # možeš promijeniti na 3, 10, što želiš
        )
    except Exception as e:
        st.warning(f"⚠ Ne mogu izračunati expected lineup strength (player stats). Detalji: {e}")
        for col in ["exp_strength_home", "exp_strength_away", "exp_strength_diff"]:
            if col not in preds.columns:
                preds[col] = np.nan

    if "xg_home" in preds.columns and "xg_away" in preds.columns:
        preds["xg_diff"] = preds["xg_home"] - preds["xg_away"]
        preds["xg_total"] = preds["xg_home"] + preds["xg_away"]
    else:
        preds["xg_diff"] = np.nan
        preds["xg_total"] = np.nan

    played_ft = preds[(preds["is_fixture"] == False) & (preds["actual"].notna())].copy()

    if not played_ft.empty:
        played_ft["poisson_pick"] = played_ft[["p_home", "p_draw", "p_away"]].idxmax(axis=1).map(
            {"p_home": "H", "p_draw": "D", "p_away": "A"}
        )
        played_ft["ai_pick"] = played_ft[["ai_p_home", "ai_p_draw", "ai_p_away"]].idxmax(axis=1).map(
            {"ai_p_home": "H", "ai_p_draw": "D", "ai_p_away": "A"}
        )
        played_ft["hit_poisson"] = (played_ft["poisson_pick"] == played_ft["actual"]).astype(int)
        played_ft["hit_ai"] = (played_ft["ai_pick"] == played_ft["actual"]).astype(int)
        acc_poi_ft = played_ft["hit_poisson"].mean()
        acc_ai_ft = played_ft["hit_ai"].mean()
    else:
        acc_poi_ft = np.nan
        acc_ai_ft = np.nan

    played_goals = preds[(preds["is_fixture"] == False) & (preds["actual_over25"].notna())].copy()
    if not played_goals.empty and "ai_p_over25" in played_goals.columns:
        played_goals["poi_over25_pick"] = (played_goals["p_over25_poi"] >= 0.5).astype(int)
        played_goals["ai_over25_pick"] = (played_goals["ai_p_over25"] >= 0.5).astype(int)
        played_goals["hit_poi_over25"] = (played_goals["poi_over25_pick"] == played_goals["actual_over25"]).astype(int)
        played_goals["hit_ai_over25"] = (played_goals["ai_over25_pick"] == played_goals["actual_over25"]).astype(int)
        acc_poi_ou = played_goals["hit_poi_over25"].mean()
        acc_ai_ou = played_goals["hit_ai_over25"].mean()

        played_btts = played_goals[played_goals["actual_btts"].notna()].copy()
        if not played_btts.empty and "ai_p_btts_yes" in played_btts.columns:
            played_btts["poi_btts_pick"] = (played_btts["p_btts_poi"] >= 0.5).astype(int)
            played_btts["ai_btts_pick"] = (played_btts["ai_p_btts_yes"] >= 0.5).astype(int)
            played_btts["hit_poi_btts"] = (played_btts["poi_btts_pick"] == played_btts["actual_btts"]).astype(int)
            played_btts["hit_ai_btts"] = (played_btts["ai_btts_pick"] == played_btts["actual_btts"]).astype(int)
            acc_poi_btts = played_btts["hit_poi_btts"].mean()
            acc_ai_btts = played_btts["hit_ai_btts"].mean()
        else:
            acc_poi_btts = np.nan
            acc_ai_btts = np.nan
    else:
        acc_poi_ou = np.nan
        acc_ai_ou = np.nan
        acc_poi_btts = np.nan
        acc_ai_btts = np.nan

    if not played_goals.empty:
        roi_poi_ou, n_poi_ou = compute_roi_binary(
            played_goals, prob_col="p_over25_poi", odds_col="book_over25",
            actual_col="actual_over25", edge_threshold=0.0
        )
        roi_ai_ou, n_ai_ou = compute_roi_binary(
            played_goals, prob_col="ai_p_over25", odds_col="book_over25",
            actual_col="actual_over25", edge_threshold=0.0
        )
        roi_poi_btts, n_poi_btts = compute_roi_binary(
            played_goals, prob_col="p_btts_poi", odds_col="book_btts_yes",
            actual_col="actual_btts", edge_threshold=0.0
        )
        roi_ai_btts, n_ai_btts = compute_roi_binary(
            played_goals, prob_col="ai_p_btts_yes", odds_col="book_btts_yes",
            actual_col="actual_btts", edge_threshold=0.0
        )
    else:
        roi_poi_ou = roi_ai_ou = roi_poi_btts = roi_ai_btts = np.nan
        n_poi_ou = n_ai_ou = n_poi_btts = n_ai_btts = 0

    num_leagues = len(preds["league"].unique())
    total_matches = len(preds)
    fixtures_count = preds[preds["is_fixture"] == True].shape[0]

    acc_ai_ft_str = f"{acc_ai_ft:.1%}" if not np.isnan(acc_ai_ft) else "N/A"
    acc_ai_ou_str = f"{acc_ai_ou:.1%}" if not np.isnan(acc_ai_ou) else "N/A"
    acc_ai_btts_str = f"{acc_ai_btts:.1%}" if not np.isnan(acc_ai_btts) else "N/A"

    roi_ai_ou_str = f"{roi_ai_ou*100:.1f}%" if not np.isnan(roi_ai_ou) else "N/A"
    roi_ai_btts_str = f"{roi_ai_btts*100:.1f}%" if not np.isnan(roi_ai_btts) else "N/A"

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(
            f"""
            <div class="kpi-card">
              <div class="kpi-label">Leagues</div>
              <div class="kpi-main">{num_leagues}</div>
              <div class="kpi-sub">Active in the model</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with c2:
        st.markdown(
            f"""
            <div class="kpi-card">
              <div class="kpi-label">Matches (history + fixtures)</div>
              <div class="kpi-main">{total_matches}</div>
              <div class="kpi-sub">Total matches in dataset</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with c3:
        st.markdown(
            f"""
            <div class="kpi-card">
              <div class="kpi-label">Fixtures (upcoming)</div>
              <div class="kpi-main">{fixtures_count}</div>
              <div class="kpi-sub">Ready for betting</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with c4:
        st.markdown(
            f"""
            <div class="kpi-card">
              <div class="kpi-label">AI FT 1X2 accuracy</div>
              <div class="kpi-main">{acc_ai_ft_str}</div>
              <div class="kpi-sub">On played matches</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    g1, g2, g3, g4 = st.columns(4)
    with g1:
        st.markdown(
            f"""
            <div class="kpi-card">
              <div class="kpi-label">AI OU 2.5 accuracy</div>
              <div class="kpi-main">{acc_ai_ou_str}</div>
              <div class="kpi-sub">Over/Under 2.5 goals</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with g2:
        st.markdown(
            f"""
            <div class="kpi-card">
              <div class="kpi-label">AI BTTS accuracy</div>
              <div class="kpi-main">{acc_ai_btts_str}</div>
              <div class="kpi-sub">Both Teams To Score</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with g3:
        st.markdown(
            f"""
            <div class="kpi-card">
              <div class="kpi-label">AI OU 2.5 ROI (flat stake)</div>
              <div class="kpi-main">{roi_ai_ou_str}</div>
              <div class="kpi-sub">Conceptual backtest, no fees</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with g4:
        st.markdown(
            f"""
            <div class="kpi-card">
              <div class="kpi-label">AI BTTS ROI (flat stake)</div>
              <div class="kpi-main">{roi_ai_btts_str}</div>
              <div class="kpi-sub">Conceptual backtest, no fees</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown("---")

    tab0, tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        "🔑 Landing & Pricing",
        "🏠 Overview",
        "📊 xG Analysis",
        "🎯 FT 1X2 (details)",
        "🥅 Goals OU/BTTS (details)",
        "🔮 Fixtures & value bets",
        "📥 Excel export",
    ])

    # TAB 0 – Landing
    with tab0:
        render_landing_page()

    # TAB 1 – Overview
    with tab1:
        st.subheader("Model overview – FT 1X2 and goals")

        c1_over, c2_over = st.columns(2)
        with c1_over:
            st.markdown("#### FT 1X2 – Poisson vs AI")
            if played_ft.empty:
                st.info("No played matches available for FT 1X2 analysis.")
            else:
                st.write(f"**Poisson FT 1X2 accuracy:** {acc_poi_ft:.1%}")
                st.write(f"**AI FT 1X2 accuracy:** {acc_ai_ft:.1%}")
                same_pick_ratio = (played_ft["poisson_pick"] == played_ft["ai_pick"]).mean()
                st.caption(f"Share of matches where Poisson and AI give the same pick: {same_pick_ratio:.1%}")

        with c2_over:
            st.markdown("#### Goals – OU 2.5 & BTTS")
            if np.isnan(acc_ai_ou):
                st.info("Not enough data for goals analysis.")
            else:
                st.write(f"**OU 2.5 – Poisson:** {acc_poi_ou:.1%} | **AI:** {acc_ai_ou:.1%}")
                if not np.isnan(acc_poi_btts) and not np.isnan(acc_ai_btts):
                    st.write(f"**BTTS – Poisson:** {acc_poi_btts:.1%} | **AI:** {acc_ai_btts:.1%}")
                elif not np.isnan(acc_poi_btts):
                    st.write(f"**BTTS – Poisson:** {acc_poi_btts:.1%} | **AI:** N/A")
                elif not np.isnan(acc_ai_btts):
                    st.write(f"**BTTS – Poisson:** N/A | **AI:** {acc_ai_btts:.1%}")
                else:
                    st.write("**BTTS – Poisson:** N/A | **AI:** N/A")

                if not np.isnan(roi_ai_ou):
                    st.write(f"**AI OU 2.5 ROI (flat 1u, edge>0):** {roi_ai_ou*100:.1f}% (bets: {n_ai_ou})")
                if not np.isnan(roi_ai_btts):
                    st.write(f"**AI BTTS ROI (flat 1u, edge>0):** {roi_ai_btts*100:.1f}% (bets: {n_ai_btts})")

        st.markdown("#### λ distribution (expected goals)")
        cc1, cc2 = st.columns(2)
        with cc1:
            st.caption("λ home")
            st.bar_chart(preds["lambda_home"])
        with cc2:
            st.caption("λ away")
            st.bar_chart(preds["lambda_away"])

    # TAB 2 – xG Analysis
    with tab2:
        st.subheader("📊 xG Analysis – post-match xG + model λ")

        played_with_xg = preds[
            (preds["is_fixture"] == False) &
            (preds["xg_home"].notna()) &
            (preds["xg_away"].notna())
        ].copy() if "xg_home" in preds.columns else pd.DataFrame()

        if played_with_xg.empty:
            st.info("No xG data mapped – check team_mapping.xlsx i xg_*.xlsx.")
        else:
            played_with_xg["total_xg"] = played_with_xg["xg_home"] + played_with_xg["xg_away"]
            played_with_xg["xg_diff_abs"] = (played_with_xg["xg_home"] - played_with_xg["xg_away"]).abs()

            c1x, c2x, c3x = st.columns(3)
            with c1x:
                st.metric("Matches with xG", f"{played_with_xg.shape[0]}")
            with c2x:
                st.metric("Avg total xG", f"{played_with_xg['total_xg'].mean():.2f}")
            with c3x:
                st.metric("Avg |xG diff|", f"{played_with_xg['xg_diff_abs'].mean():.2f}")

            st.markdown("#### Top 30 matches by xG dominance")
            top_dom = played_with_xg.sort_values("xg_diff_abs", ascending=False).head(30)
            cols_show = [
                "league", "match_date", "home", "away",
                "xg_home", "xg_away", "xg_diff",
                "lambda_home", "lambda_away",
                "p_home", "p_draw", "p_away",
                "actual"
            ]
            cols_show = [c for c in cols_show if c in top_dom.columns]
            st.dataframe(top_dom[cols_show].round(3), use_container_width=True)

            st.markdown("#### xG diff over time")
            tmp = played_with_xg.dropna(subset=["match_date"]).copy()
            tmp = tmp.sort_values("match_date")
            st.line_chart(tmp.set_index("match_date")["xg_diff"])

        st.markdown("---")
        st.subheader("📦 Raw xG cache (svi redovi iz xg_*.xlsx)")

        xg_df = load_xg_cache()
        if xg_df.empty:
            st.info("No xG data loaded – check xg_*.xlsx in data/api_football.")
        else:
            st.dataframe(xg_df.head(200), use_container_width=True)

    # TAB 3 – FT 1X2 details
    with tab3:
        st.subheader("FT 1X2 – detailed Poisson vs AI comparison")

        if played_ft.empty:
            st.info("No played matches available.")
        else:
            show_cols = [
                "league", "match_date", "home", "away",
                "actual", "poisson_pick", "ai_pick",
                "hit_poisson", "hit_ai",
                "p_home", "p_draw", "p_away",
                "ai_p_home", "ai_p_draw", "ai_p_away",
                "book_home", "book_draw", "book_away",
                "edge_home", "edge_draw", "edge_away",
                "kelly_home", "kelly_draw", "kelly_away",
            ]
            show_cols = [c for c in show_cols if c in played_ft.columns]
            st.dataframe(played_ft[show_cols].round(3), use_container_width=True)

    # TAB 4 – Goals details
    with tab4:
        st.subheader("Goals – OU 1.5 / 2.5 & BTTS (Poisson vs AI + ROI)")

        if played_goals.empty or "ai_p_over25" not in played_goals.columns:
            st.info("Not enough data for goals analysis.")
        else:
            c1_goals, c2_goals = st.columns(2)
            c1_goals.metric("OU 2.5 – Poisson accuracy", f"{acc_poi_ou:.1%}")
            c2_goals.metric("OU 2.5 – AI accuracy", f"{acc_ai_ou:.1%}")

            c3_goals, c4_goals = st.columns(2)
            if not np.isnan(acc_poi_btts):
                c3_goals.metric("BTTS – Poisson accuracy", f"{acc_poi_btts:.1%}")
                c4_goals.metric("BTTS – AI accuracy", f"{acc_ai_btts:.1%}")
            else:
                c3_goals.metric("BTTS – Poisson accuracy", "N/A")
                c4_goals.metric("BTTS – AI accuracy", "N/A")

            st.markdown("#### ROI simulation (flat 1u stake, bet where p·odds - 1 > 0)")

            r1, r2 = st.columns(2)
            if not np.isnan(roi_poi_ou):
                r1.metric("OU 2.5 – Poisson ROI", f"{roi_poi_ou*100:.1f}%", help=f"#bets: {n_poi_ou}")
            else:
                r1.metric("OU 2.5 – Poisson ROI", "N/A")

            if not np.isnan(roi_ai_ou):
                r2.metric("OU 2.5 – AI ROI", f"{roi_ai_ou*100:.1f}%", help=f"#bets: {n_ai_ou}")
            else:
                r2.metric("OU 2.5 – AI ROI", "N/A")

            r3, r4 = st.columns(2)
            if not np.isnan(roi_poi_btts):
                r3.metric("BTTS – Poisson ROI", f"{roi_poi_btts*100:.1f}%", help=f"#bets: {n_poi_btts}")
            else:
                r3.metric("BTTS – Poisson ROI", "N/A")

            if not np.isnan(roi_ai_btts):
                r4.metric("BTTS – AI ROI", f"{roi_ai_btts*100:.1f}%", help=f"#bets: {n_ai_btts}")
            else:
                r4.metric("BTTS – AI ROI", "N/A")

            st.markdown("#### Detailed table (played matches)")
            cols_show = [
                "league", "match_date", "home", "away",
                "actual_over25", "poi_over25_pick", "ai_over25_pick",
                "hit_poi_over25", "hit_ai_over25",
                "p_over15_poi", "p_over25_poi", "ai_p_over25", "book_over15", "book_over25",
                "actual_btts", "p_btts_poi", "ai_p_btts_yes", "book_btts_yes", "book_btts_no",
                "ai_total_goals",
            ]
            cols_show = [c for c in cols_show if c in played_goals.columns]
            st.dataframe(played_goals[cols_show].round(3), use_container_width=True)

    # TAB 5 – Fixtures & value bets
    with tab5:
        st.subheader("🔮 Fixtures – FT 1X2 + Goals value bets")

        fixtures = preds[preds["is_fixture"] == True].copy()
        st.caption(f"Raw fixtures from web: {raw_fixtures_count} | Fixtures with model λ: {fixtures.shape[0]}")

        if fixtures.empty:
            st.warning("No fixture predictions to display.")
        else:
            fixtures = finalize_picks(fixtures)
            fixtures = apply_betting_rules(fixtures, bet_settings)
            md_series = pd.to_datetime(fixtures["match_date"], errors="coerce")
            min_md = md_series.min()
            max_md = md_series.max()

            if pd.isna(min_md) or pd.isna(max_md):
                default_range = None
            else:
                default_range = (min_md.date(), max_md.date())

            st.markdown("<div class='filter-bar'>", unsafe_allow_html=True)
            f1, f2, f3 = st.columns([1.2, 1, 1.2])
            with f1:
                league_filter = st.selectbox(
                    "League filter",
                    ["All"] + sorted(fixtures["league"].dropna().unique().tolist()),
                    index=0,
                )
            with f2:
                ui_min_edge = st.slider("Min edge (all markets)", 0.0, 0.30, 0.02, 0.01)
            with f3:
                ui_min_kelly = st.slider("Min Kelly (all markets)", 0.0, 0.15, 0.01, 0.005)
            s1, s2 = st.columns([1.3, 1])
            with s1:
                team_search = st.text_input("Search team (home/away)", "")
            with s2:
                if default_range:
                    date_range = st.date_input(
                        "Date range (optional)",
                        value=default_range,
                        min_value=default_range[0],
                        max_value=default_range[1],
                        help="Choose start and end date"
                    )
                else:
                    date_range = None

            st.markdown("</div>", unsafe_allow_html=True)

            f = fixtures.copy()

            if league_filter != "All":
                f = f[f["league"] == league_filter]

            if team_search:
                ts = team_search.lower()
                f = f[
                    f["home"].str.lower().str.contains(ts)
                    | f["away"].str.lower().str.contains(ts)
                ]

            if isinstance(date_range, (list, tuple)) and len(date_range) == 2:
                start_d, end_d = date_range
                if start_d and end_d:
                    md = pd.to_datetime(f["match_date"], errors="coerce").dt.date
                    f = f[(md >= start_d) & (md <= end_d)]
            show_valid_only = st.checkbox("Hide HARD FAIL", True)
            hide_soft = st.checkbox("Hide SOFT FLAG (huge-value / low-consensus)", True)

            if show_valid_only and "hard_fail" in f.columns:
                f = f[~f["hard_fail"]]
            if hide_soft and "soft_flag" in f.columns:
                f = f[~f["soft_flag"]]

            edge_cut = max(min_edge, ui_min_edge)
            kelly_cut = max(min_kelly, ui_min_kelly)

            st.markdown("#### 🎴 Match cards (premium view)")

            kelly_cols_all = [c for c in f.columns if c.startswith("kelly_")]
            if kelly_cols_all:
                f["max_kelly_any"] = f[kelly_cols_all].max(axis=1)
                f_cards = f.sort_values("max_kelly_any", ascending=False).head(40)
            else:
                f_cards = f.copy().head(40)

            if f_cards.empty:
                st.info("No fixtures for current filters.")
            else:
                for _, r in f_cards.iterrows():

                    # ---- VALUE BADGE (edge) ----
                    edge_candidates = [
                        r.get("edge_home", np.nan),
                        r.get("edge_draw", np.nan),
                        r.get("edge_away", np.nan),
                        r.get("edge_ou15_poi", np.nan),
                        r.get("edge_ou25_ai", np.nan),
                        r.get("edge_btts_ai", np.nan),
                    ]
                    best_edge = np.nanmax(edge_candidates)
                    badge_html = ""
                    if not np.isnan(best_edge) and best_edge > 0:
                        badge_html = f'<span class="value-badge">VALUE +{best_edge * 100:.1f}%</span>'

                    # ---- RISK BADGE ----
                    rl = str(r.get("display_risk", r.get("risk_level", "NONE"))).upper()
                    risk_html = ""
                    if rl == "HIGH":
                        risk_html = '<span class="risk-badge-high">HIGH</span>'
                    elif rl == "MEDIUM":
                        risk_html = '<span class="risk-badge-medium">MEDIUM</span>'
                    elif rl == "LOW":
                        risk_html = '<span class="risk-badge-low">LOW</span>'

                    # ---- BASIC FIELDS ----
                    match_dt = r.get("match_date", None)
                    match_dt_str = str(match_dt) if pd.notna(match_dt) else ""

                    ai_total = r.get("ai_total_goals", np.nan)
                    ai_total_str = f"{ai_total:.2f}" if pd.notna(ai_total) else "-"

                    book_home = r.get("book_home", np.nan)
                    book_draw = r.get("book_draw", np.nan)
                    book_away = r.get("book_away", np.nan)

                    book_home_str = f"{book_home:.2f}" if pd.notna(book_home) else "-"
                    book_draw_str = f"{book_draw:.2f}" if pd.notna(book_draw) else "-"
                    book_away_str = f"{book_away:.2f}" if pd.notna(book_away) else "-"

                    ai_over = r.get("ai_p_over25", np.nan)
                    ai_over_str = f"{ai_over:.0%}" if pd.notna(ai_over) else "N/A"

                    ai_btts = r.get("ai_p_btts_yes", np.nan)
                    ai_btts_str = f"{ai_btts:.0%}" if pd.notna(ai_btts) else "N/A"

                    # ---- EXPECTED LINEUP STRENGTH ----
                    exp_h = r.get("exp_strength_home", np.nan)
                    exp_a = r.get("exp_strength_away", np.nan)
                    exp_d = r.get("exp_strength_diff", np.nan)
                    if pd.notna(exp_h) and pd.notna(exp_a) and pd.notna(exp_d):
                        exp_lineup_str = f"H {exp_h:.2f} / A {exp_a:.2f} (Δ {exp_d:+.2f})"
                    else:
                        exp_lineup_str = "N/A"

                    # ---- AUTO + STAKE ----
                    auto = (r.get("auto_action", "NO BET") or "NO BET")
                    stake = float(r.get("stake_eur", 0.0) or 0.0)

                    # ---- Recommended (escape to avoid breaking HTML) ----
                    rec = r.get("display_bet", r.get("recommended_bet", "No bet"))
                    rec_safe = html.escape(str(rec))

                    # ✅ KLJUČNO: sub_html se gradi u 1 liniji (nema indentiranih "<br/>" pa nema code-blocka)
                    sub_html = (
                        f"{html.escape(str(r['league']))} — {html.escape(match_dt_str)}<br/>"
                        f"<b>Recommended:</b> {rec_safe} {badge_html} {risk_html}<br/>"
                        f"<b>Auto:</b> {html.escape(str(auto))} • <b>Stake:</b> €{stake:.2f}"
                    )

                    st.markdown(f"""
            <div class="match-card">
              <div class="match-header">{html.escape(str(r['home']))} vs {html.escape(str(r['away']))}</div>
              <div class="match-sub">{sub_html}</div>
              <div class="match-row">
                <div class="match-col">
                  <b>λ (pre-match xG):</b><br/>
                  H {r['lambda_home']:.2f} — A {r['lambda_away']:.2f}<br/>
                  Total {r['xg_pre_total']:.2f}<br/>
                  <b>Lineup strength:</b> {html.escape(exp_lineup_str)}
                </div>
                <div class="match-col">
                  <b>AI FT 1X2:</b><br/>
                  H {r['ai_p_home']:.0%} | D {r['ai_p_draw']:.0%} | A {r['ai_p_away']:.0%}<br/>
                  Total goals (AI): {ai_total_str}
                </div>
                <div class="match-col">
                  <b>Odds 1X2:</b><br/>
                  H {book_home_str} | D {book_draw_str} | A {book_away_str}
                </div>
                <div class="match-col">
                  <b>Goals / BTTS:</b><br/>
                  OU1.5 p_Poi {r['p_over15_poi']:.0%}, OU2.5 p_AI {ai_over_str}<br/>
                  BTTS p_AI {ai_btts_str}
                </div>
              </div>
            </div>
            """, unsafe_allow_html=True)

            st.markdown("---")
            st.markdown("#### 📋 Fixtures table (after filters)")
            cols_fix = [
                "league", "match_date", "home", "away",

                "lambda_home", "lambda_away",
                "xg_pre_home", "xg_pre_away", "xg_pre_total",

                "exp_strength_home", "exp_strength_away", "exp_strength_diff",

                "p_home", "p_draw", "p_away",
                "ai_p_home", "ai_p_draw", "ai_p_away",
                "p_over15_poi", "p_over25_poi", "p_btts_poi", "p_btts_no_poi",
                "ai_p_over25", "ai_p_btts_yes", "ai_total_goals",
                "book_home", "book_draw", "book_away",
                "book_over15", "book_over25",
                "book_btts_yes", "book_btts_no",
                "edge_home", "edge_draw", "edge_away",
                "kelly_home", "kelly_draw", "kelly_away",
                "edge_ou15_poi", "kelly_ou15_poi",
                "edge_ou25_poi", "kelly_ou25_poi",
                "edge_ou25_ai", "kelly_ou25_ai",
                "edge_btts_poi", "kelly_btts_poi",
                "edge_btts_ai", "kelly_btts_ai",
                "recommended_bet", "risk_level",
                "display_bet", "display_risk", "hard_fail", "soft_flag", "reasons",
            ]
            cols_fix = [c for c in cols_fix if c in f.columns]
            st.dataframe(f[cols_fix].round(3), use_container_width=True)

            st.markdown("---")
            st.subheader("⭐ Top value bets – FT 1X2")

            rows_1x2 = []
            for _, r in f.iterrows():
                for sel, p_col, odd_col, edge_col, kelly_col in [
                    ("H", "p_home", "book_home", "edge_home", "kelly_home"),
                    ("D", "p_draw", "book_draw", "edge_draw", "kelly_draw"),
                    ("A", "p_away", "book_away", "edge_away", "kelly_away"),
                ]:
                    p = r.get(p_col)
                    o = r.get(odd_col)
                    edge = r.get(edge_col)
                    kelly = r.get(kelly_col)
                    if (
                        p is not None and not np.isnan(p) and
                        o is not None and not np.isnan(o) and
                        edge is not None and not np.isnan(edge) and edge >= edge_cut and
                        kelly is not None and kelly >= kelly_cut
                    ):
                        rows_1x2.append({
                            "league": r["league"],
                            "match_date": r["match_date"],
                            "home": r["home"],
                            "away": r["away"],
                            "selection": sel,
                            "prob": p,
                            "odds": o,
                            "edge": edge,
                            "kelly": kelly,
                        })

            if rows_1x2:
                df_val_1x2 = pd.DataFrame(rows_1x2).sort_values("edge", ascending=False)
                st.dataframe(df_val_1x2.round(3), use_container_width=True)
            else:
                st.info("No FT 1X2 value bets for the current filters.")

            st.markdown("---")
            st.subheader("🔥 Top value bets – Over 1.5 (Poisson)")

            rows_ou15 = []
            for _, r in f.iterrows():
                p = r.get("p_over15_poi")
                o = r.get("book_over15")
                edge = r.get("edge_ou15_poi")
                kelly = r.get("kelly_ou15_poi")
                if (
                    p is not None and not np.isnan(p) and
                    o is not None and not np.isnan(o) and
                    edge is not None and not np.isnan(edge) and edge >= edge_cut and
                    kelly is not None and kelly >= kelly_cut
                ):
                    rows_ou15.append({
                        "league": r["league"],
                        "match_date": r["match_date"],
                        "home": r["home"],
                        "away": r["away"],
                        "selection": "Over 1.5",
                        "prob_poi": p,
                        "odds": o,
                        "edge_poi": edge,
                        "kelly_poi": kelly,
                    })
            if rows_ou15:
                df_val_ou15 = pd.DataFrame(rows_ou15).sort_values("edge_poi", ascending=False)
                st.dataframe(df_val_ou15.round(3), use_container_width=True)
            else:
                st.info("No OU 1.5 Poisson value bets or no OU 1.5 odds available for current filters.")

            st.markdown("---")
            st.subheader("🔥 Top value bets – Over 2.5 (AI)")

            rows_ou = []
            for _, r in f.iterrows():
                p = r.get("ai_p_over25")
                o = r.get("book_over25")
                edge = r.get("edge_ou25_ai")
                kelly = r.get("kelly_ou25_ai")
                if (
                    p is not None and not np.isnan(p) and
                    o is not None and not np.isnan(o) and
                    edge is not None and not np.isnan(edge) and edge >= edge_cut and
                    kelly is not None and kelly >= kelly_cut
                ):
                    rows_ou.append({
                        "league": r["league"],
                        "match_date": r["match_date"],
                        "home": r["home"],
                        "away": r["away"],
                        "selection": "Over 2.5",
                        "prob_ai": p,
                        "odds": o,
                        "edge_ai": edge,
                        "kelly_ai": kelly,
                    })
            if rows_ou:
                df_val_ou = pd.DataFrame(rows_ou).sort_values("edge_ai", ascending=False)
                st.dataframe(df_val_ou.round(3), use_container_width=True)
            else:
                st.info("No OU 2.5 AI value bets or no OU odds available for current filters.")

            st.markdown("---")
            st.subheader("💥 Top value bets – BTTS YES (AI)")

            rows_btts = []
            for _, r in f.iterrows():
                p = r.get("ai_p_btts_yes")
                o = r.get("book_btts_yes")
                edge = r.get("edge_btts_ai")
                kelly = r.get("kelly_btts_ai")
                if (
                    p is not None and not np.isnan(p) and
                    o is not None and not np.isnan(o) and
                    edge is not None and not np.isnan(edge) and edge >= edge_cut and
                    kelly is not None and kelly >= kelly_cut
                ):
                    rows_btts.append({
                        "league": r["league"],
                        "match_date": r["match_date"],
                        "home": r["home"],
                        "away": r["away"],
                        "selection": "BTTS YES",
                        "prob_ai": p,
                        "odds": o,
                        "edge_ai": edge,
                        "kelly_ai": kelly,
                    })
            if rows_btts:
                df_val_btts = pd.DataFrame(rows_btts).sort_values("edge_ai", ascending=False)
                st.dataframe(df_val_btts.round(3), use_container_width=True)
            else:
                st.info("No BTTS AI value bets or no BTTS odds available for current filters.")

    # TAB 6 – Excel export
    with tab6:
        st.subheader("Excel export – all data + PRO fixtures")

        played = preds[preds["is_fixture"] == False].copy()
        fixtures = preds[preds["is_fixture"] == True].copy()

        fixtures_dashboard = pd.DataFrame()
        if not fixtures.empty:
            cols_fix = [
                "league", "match_date", "home", "away",

                "lambda_home", "lambda_away",
                "xg_pre_home", "xg_pre_away", "xg_pre_total",

                "exp_strength_home", "exp_strength_away", "exp_strength_diff",

                "p_home", "p_draw", "p_away",
                "ai_p_home", "ai_p_draw", "ai_p_away",

                "p_over15_poi", "p_over25_poi", "p_btts_poi", "p_btts_no_poi",
                "ai_p_over25", "ai_p_btts_yes",
                "ai_total_goals",

                "book_home", "book_draw", "book_away",
                "book_over15", "book_over25",
                "book_btts_yes", "book_btts_no",

                "edge_home", "edge_draw", "edge_away",
                "kelly_home", "kelly_draw", "kelly_away",
                "edge_ou15_poi", "kelly_ou15_poi",
                "edge_ou25_poi", "kelly_ou25_poi",
                "edge_ou25_ai", "kelly_ou25_ai",
                "edge_btts_poi", "kelly_btts_poi",
                "edge_btts_ai", "kelly_btts_ai",
            ]

            cols_fix = [c for c in cols_fix if c in fixtures.columns]
            fixtures_dashboard = fixtures[cols_fix].copy()
            prob_cols = [c for c in fixtures_dashboard.columns if c.startswith(("p_", "ai_p_"))]
            fixtures_dashboard[prob_cols] = fixtures_dashboard[prob_cols].round(3)

        buffer_all = BytesIO()
        with pd.ExcelWriter(buffer_all, engine="openpyxl") as writer:
            preds.to_excel(writer, index=False, sheet_name="Predictions_all")
            if not played.empty:
                played.to_excel(writer, index=False, sheet_name="Played_raw")
            if not fixtures.empty:
                fixtures.to_excel(writer, index=False, sheet_name="Fixtures_raw")
            if not fixtures_dashboard.empty:
                fixtures_dashboard.to_excel(writer, index=False, sheet_name="Fixtures_dashboard")
        buffer_all.seek(0)

        st.download_button(
            label="📥 Download FULL Excel (all tables, incl xG where available)",
            data=buffer_all,
            file_name=f"poisson_ai_xg_full_{season}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.markdown("---")
        st.subheader("📥 PRO Excel – Fixtures only (mini app for clients)")

        if fixtures_dashboard.empty:
            st.warning("No fixtures data available for PRO export.")
        else:
            buffer_fix = build_pro_fixtures_excel(fixtures_dashboard, season)
            st.download_button(
                label="📥 Download Fixtures PRO Excel",
                data=buffer_fix,
                file_name=f"fixtures_PRO_{season}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    st.markdown(
        """
        <div class="gm-footer">
            © <span>2025</span> GOALMIND PRO • Advanced football prediction engine (Poisson + AI + xG + Kelly).
        </div>
        """,
        unsafe_allow_html=True,
    )


if __name__ == "__main__":
    if check_password():
        main()
